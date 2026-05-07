"""run5 — Old Jama vs New Jama, with Name + Description similarity.

Same task and structure as run4: pair each Old row with its best-matching New
row, classify identical/mismatch, and list unclaimed New rows at the bottom.
What's new in run5:

  * Description similarity is computed alongside Name similarity. Matching
    stays Name-based (Names are the reliable signal — Descriptions are sparse),
    but both similarity scores show up in the output as separate columns so
    the customer can spot 'name same, description changed' pairs.
  * The LLM verification prompt now includes both Names AND Descriptions, so
    the model has richer context for the identical/mismatch call.
  * Auto-identical fast path now requires Name similarity >= 0.97 AND
    Description similarity >= 0.97 (or one side empty), so we don't auto-pass
    a row whose Name is unchanged but whose substantive content evolved.

Status vocabulary unchanged:
  identical   — same underlying requirement
  mismatch    — paired by Name similarity but the LLM says they're different
  missing     — Old row has no good New candidate
  unmatched   — New row not claimed by any Old

Usage:
  python run5.py path/to/input.xlsx
  python run5.py path/to/input.xlsx --out path/to/output.xlsx --limit 30
"""

import argparse
import asyncio
import json
import re
import sys
from dataclasses import asdict, dataclass
from pathlib import Path

import httpx
import numpy as np
from openai import AsyncOpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sentence_transformers import SentenceTransformer
from tqdm import tqdm

import extract_excel


# ---------- config ----------

DEFAULT_LLM_URL = "http://localhost:8000/v1"
DEFAULT_LLM_MODEL = "Qwen/Qwen3.6-27B"
DEFAULT_EMBED_MODEL = "BAAI/bge-small-en-v1.5"
DEFAULT_CONCURRENCY = 16
DEFAULT_LLM_TIMEOUT = 180.0

OLD_SHEET_DEFAULT = "Old Jama"
NEW_SHEET_DEFAULT = "New Jama"

# Below this cosine similarity an Old gets no candidate at all.
DEFAULT_MATCH_THRESHOLD = 0.40
# Above this, skip the LLM and auto-mark identical (the names are essentially
# the same string).
AUTO_IDENTICAL_THRESHOLD = 0.97

# Column-name candidates per side. We accept light variations in casing/spacing.
ID_COL_CANDIDATES = ["Project ID", "ID", "Global ID", "GID", "Jama ID", "Item ID"]
NAME_COL_CANDIDATES = ["Name", "Title"]
TYPE_COL_CANDIDATES = ["Item Type", "Type", "Category"]
DESC_COL_CANDIDATES = ["Description", "Text", "Requirement"]

# Status vocabulary written into the Status column.
ST_IDENTICAL = "identical"
ST_MISMATCH = "mismatch"
ST_MISSING = "missing"      # Old row, no matching New
ST_UNMATCHED = "unmatched"  # New row, no matching Old
ST_ERROR = "error"

STATUS_FILLS = {
    ST_IDENTICAL: PatternFill("solid", fgColor="C6EFCE"),  # green
    ST_MISMATCH:  PatternFill("solid", fgColor="FFEB9C"),  # yellow
    ST_MISSING:   PatternFill("solid", fgColor="FFC7CE"),  # red
    ST_UNMATCHED: PatternFill("solid", fgColor="DCE6F1"),  # light blue
    ST_ERROR:     PatternFill("solid", fgColor="F4B084"),  # orange
}


# ---------- data ----------

@dataclass
class Row:
    """A single row from either Old Jama or New Jama."""
    sheet_index: int  # 0-based position within its sheet (preserves input order)
    jama_id: str
    item_type: str
    name: str
    description: str
    hyperlink: str | None = None  # extracted from the input xlsx, applied to the ID cell


@dataclass
class Verdict:
    """Result of an LLM comparison for one (old, new) pair."""
    old_index: int
    new_index: int
    status: str            # identical / mismatch / error
    similarity: float
    reasoning: str


# ---------- helpers ----------

def first_key(sample: dict, candidates: list[str]) -> str | None:
    for c in candidates:
        if c in sample:
            return c
    lower = {k.lower(): k for k in sample}
    for c in candidates:
        if c.lower() in lower:
            return lower[c.lower()]
    return None


def load_rows(xlsx_path: Path, sheet: str | int, work_json: Path,
              re_extract: bool) -> list[Row]:
    """Run extract_excel for the given sheet and load the resulting JSON into Row objects."""
    if re_extract or not work_json.exists():
        print(f"extracting sheet {sheet!r} from {xlsx_path.name} -> {work_json.name} ...")
        extract_excel.extract(xlsx_path, sheet=sheet, out_path=work_json, header=0)
    else:
        print(f"reusing existing {work_json.name}")

    data = json.loads(work_json.read_text())
    rows = data["rows"]
    if not rows:
        return []

    sample = rows[0]
    id_key = first_key(sample, ID_COL_CANDIDATES) or "Project ID"
    name_key = first_key(sample, NAME_COL_CANDIDATES) or "Name"
    type_key = first_key(sample, TYPE_COL_CANDIDATES) or "Item Type"
    desc_key = first_key(sample, DESC_COL_CANDIDATES) or "Description"

    out: list[Row] = []
    linked = 0
    for i, r in enumerate(rows):
        jama_id = str(r.get(id_key, "")).strip()
        name = str(r.get(name_key, "")).strip()
        itype = str(r.get(type_key, "")).strip()
        desc = str(r.get(desc_key, "")).strip()
        if not (jama_id or name):
            continue  # blank row
        # Hyperlinks are stored by extract_excel.py keyed by column name. Try
        # the ID column first; fall back to the first hyperlink in the row
        # (sometimes Jama exports put the URL on a different column).
        links = r.get("_hyperlinks") or {}
        hyperlink = links.get(id_key)
        if not hyperlink and links:
            hyperlink = next(iter(links.values()), None)
        if hyperlink:
            linked += 1
        out.append(Row(
            sheet_index=i, jama_id=jama_id, item_type=itype,
            name=name, description=desc, hyperlink=hyperlink,
        ))
    print(f"  {len(out)} usable rows from sheet {sheet!r} "
          f"(keys: id={id_key!r}, name={name_key!r}, type={type_key!r}, desc={desc_key!r}; "
          f"{linked} with hyperlinks)")
    return out


# ---------- embedding + matching ----------

def embed(model: SentenceTransformer, texts: list[str], label: str) -> np.ndarray:
    if not texts:
        return np.zeros((0, model.get_sentence_embedding_dimension()), dtype=np.float32)
    print(f"embedding {len(texts)} {label} on CPU...")
    vecs = model.encode(
        texts, batch_size=32, convert_to_numpy=True,
        normalize_embeddings=True, show_progress_bar=True,
    )
    return vecs.astype(np.float32)


def greedy_match(sim: np.ndarray, threshold: float) -> dict[int, tuple[int, float]]:
    """Greedy 1:1 bipartite assignment. Returns {old_idx: (new_idx, similarity)}.

    We sort all (old, new, sim) triples by similarity descending and claim pairs
    in that order. Each Old and each New can be claimed at most once. Pairs
    below `threshold` are ignored — the Old will be reported as 'missing'.
    """
    n_old, n_new = sim.shape
    # Skip the all-pairs explosion if either side is empty.
    if n_old == 0 or n_new == 0:
        return {}

    # argsort flat similarities once (descending). Big enough? 200x400 = 80k
    # entries — trivial.
    flat = sim.flatten()
    order = np.argsort(-flat)

    matched_old: dict[int, tuple[int, float]] = {}
    matched_new: set[int] = set()
    for k in order:
        s = float(flat[k])
        if s < threshold:
            break
        i, j = divmod(int(k), n_new)
        if i in matched_old or j in matched_new:
            continue
        matched_old[i] = (j, s)
        matched_new.add(j)
        if len(matched_old) == n_old or len(matched_new) == n_new:
            break
    return matched_old


# ---------- LLM ----------

CLASSIFY_SYSTEM = (
    "You are comparing two Jama requirement names — one from an old version of "
    "a spec, one from a new version. Decide whether they describe the SAME "
    "underlying requirement.\n\n"
    "Status (binary):\n"
    "- identical: same underlying requirement, regardless of wording, "
    "  capitalization, abbreviation, paraphrasing, or punctuation differences.\n"
    "- mismatch: different requirements, even if topically related.\n\n"
    "OUTPUT FORMAT — STRICT:\n"
    "Your entire response must be a single JSON object and nothing else.\n"
    "- Do NOT narrate. Do NOT explain your thinking outside the JSON.\n"
    "- Your first character MUST be '{' and your last character MUST be '}'.\n"
    "- Reasoning, if any, goes in the 'reasoning' field as one short sentence."
)

CLASSIFY_USER = """Old item:
ID: {old_id}
Item Type: {old_type}
Name: {old_name}
Description: {old_desc}

New item:
ID: {new_id}
Item Type: {new_type}
Name: {new_name}
Description: {new_desc}

Return JSON:
{{
  "status": "identical" | "mismatch",
  "reasoning": "<one short sentence>"
}}
"""


def parse_llm_json(raw: str) -> dict:
    s = raw.strip()
    if s.startswith("```"):
        s = re.sub(r"^```(?:json)?\s*", "", s)
        s = re.sub(r"\s*```$", "", s)
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", s, re.DOTALL)
        if not m:
            raise
        return json.loads(m.group(0))


def llm_content(msg) -> str:
    c = (msg.content or "").strip()
    if c:
        return c
    return (getattr(msg, "reasoning_content", None) or "").strip()


def _looks_like_json(text: str) -> bool:
    s = text.strip()
    if not s.startswith("{"):
        return False
    try:
        parse_llm_json(s)
        return True
    except json.JSONDecodeError:
        return False


NO_THINK_TOKEN = "\n\n/no_think"


def _disable_thinking(messages: list[dict]) -> list[dict]:
    out = [dict(m) for m in messages]
    for m in reversed(out):
        if m.get("role") == "user":
            m["content"] = (m.get("content") or "") + NO_THINK_TOKEN
            break
    return out


async def _llm_call(client: AsyncOpenAI, model: str, messages: list[dict],
                    max_tokens: int, fallback_max_tokens: int = 2000) -> tuple[str, dict]:
    """Two-attempt strategy mirroring run3: full reasoning with json_object first,
    then /no_think + json_object + bumped sampling if attempt 1 returns
    empty/non-JSON content."""
    diag: dict = {}

    # Attempt 1
    try:
        resp = await client.chat.completions.create(
            model=model, messages=messages,
            temperature=0.0, max_tokens=max_tokens,
            response_format={"type": "json_object"},
        )
    except Exception as e_rf:  # noqa: BLE001
        diag["a1_no_response_format"] = f"{type(e_rf).__name__}: {e_rf}"
        resp = await client.chat.completions.create(
            model=model, messages=messages,
            temperature=0.0, max_tokens=max_tokens,
        )
    choice = resp.choices[0]
    content = llm_content(choice.message)
    diag["a1_finish"] = choice.finish_reason
    if content and _looks_like_json(content):
        return content, diag
    if content:
        diag["a1_non_json"] = content[:200]

    # Attempt 2 — /no_think + aggressive sampling.
    fb_messages = _disable_thinking(messages)
    try:
        try:
            resp = await client.chat.completions.create(
                model=model, messages=fb_messages,
                temperature=0.7, top_p=0.9, frequency_penalty=0.5,
                max_tokens=fallback_max_tokens,
                response_format={"type": "json_object"},
            )
        except Exception as e_rf:  # noqa: BLE001
            diag["a2_no_response_format"] = f"{type(e_rf).__name__}: {e_rf}"
            resp = await client.chat.completions.create(
                model=model, messages=fb_messages,
                temperature=0.7, top_p=0.9, frequency_penalty=0.5,
                max_tokens=fallback_max_tokens,
            )
        choice = resp.choices[0]
        content = llm_content(choice.message)
        diag["a2_finish"] = choice.finish_reason
        diag["a2_no_think"] = True
        if content and _looks_like_json(content):
            return content, diag
        if content:
            diag["a2_non_json"] = content[:200]
    except Exception as e:  # noqa: BLE001
        diag["a2_error"] = f"{type(e).__name__}: {e}"
    return "", diag


async def classify_pair(client: AsyncOpenAI, model: str,
                        old: Row, new: Row, similarity: float) -> Verdict:
    user = CLASSIFY_USER.format(
        old_id=old.jama_id, old_type=old.item_type or "",
        old_name=old.name[:1000],
        old_desc=(old.description or "(empty)")[:2000],
        new_id=new.jama_id, new_type=new.item_type or "",
        new_name=new.name[:1000],
        new_desc=(new.description or "(empty)")[:2000],
    )
    messages = [
        {"role": "system", "content": CLASSIFY_SYSTEM},
        {"role": "user", "content": user},
    ]
    try:
        content, diag = await _llm_call(
            client, model, messages, max_tokens=8000, fallback_max_tokens=2000,
        )
    except Exception as e:  # noqa: BLE001
        return Verdict(
            old_index=old.sheet_index, new_index=new.sheet_index,
            status=ST_ERROR, similarity=similarity,
            reasoning=f"LLM call failed: {type(e).__name__}: {e}",
        )
    if not content:
        # Auto-default to mismatch with diag in reasoning — same pattern as run3.
        return Verdict(
            old_index=old.sheet_index, new_index=new.sheet_index,
            status=ST_MISMATCH, similarity=similarity,
            reasoning=f"(auto-default) model returned empty/non-JSON. diag={diag}",
        )
    try:
        data = parse_llm_json(content)
    except json.JSONDecodeError:
        return Verdict(
            old_index=old.sheet_index, new_index=new.sheet_index,
            status=ST_ERROR, similarity=similarity,
            reasoning=f"JSON parse failed. raw={content[:300]}",
        )
    status = str(data.get("status", "")).strip().lower()
    if status not in {ST_IDENTICAL, ST_MISMATCH}:
        return Verdict(
            old_index=old.sheet_index, new_index=new.sheet_index,
            status=ST_ERROR, similarity=similarity,
            reasoning=f"unexpected status from LLM: {status!r}",
        )
    return Verdict(
        old_index=old.sheet_index, new_index=new.sheet_index,
        status=status, similarity=similarity,
        reasoning=str(data.get("reasoning", ""))[:1000],
    )


# ---------- checkpoint ----------

def load_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        return []
    out = []
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            out.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return out


def append_jsonl(path: Path, obj: dict) -> None:
    with path.open("a") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")


# ---------- runner ----------

async def run_classify(client, model, jobs, checkpoint_path, done_keys, concurrency):
    sem = asyncio.Semaphore(concurrency)
    pbar = tqdm(total=len(jobs),
                initial=sum(1 for o, n, _ in jobs if (o.sheet_index, n.sheet_index) in done_keys),
                desc="classify", unit="pair")
    results: list[Verdict] = []

    async def worker(old: Row, new: Row, sim: float):
        key = (old.sheet_index, new.sheet_index)
        if key in done_keys:
            return
        async with sem:
            v = await classify_pair(client, model, old, new, sim)
        append_jsonl(checkpoint_path, asdict(v))
        results.append(v)
        pbar.update(1)

    await asyncio.gather(*(worker(o, n, s) for o, n, s in jobs))
    pbar.close()
    return results


# ---------- writer ----------

REPORT_COLUMNS = [
    ("Old Project ID", 18),
    ("Old Item Type", 22),
    ("Old Name", 50),
    ("Old Description", 50),
    ("New Project ID", 18),
    ("New Item Type", 22),
    ("New Name", 50),
    ("New Description", 50),
    ("Status", 14),
    ("Name Similarity", 14),
    ("Description Similarity", 18),
    ("Reasoning", 50),
]
STATUS_COL_INDEX = 9  # 1-based position of Status column (for fills)


def write_report(out_path: Path, olds: list[Row], news: list[Row],
                 matches: dict[int, tuple[int, float]],
                 verdicts: dict[tuple[int, int], Verdict],
                 desc_sim: np.ndarray) -> None:
    """Write the comparison xlsx with both Name and Description similarity scores."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Old vs New"

    # Header.
    for col_idx, (name, width) in enumerate(REPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wrap = Alignment(wrap_text=True, vertical="top")
    counts: dict[str, int] = {}
    new_claimed: set[int] = set()
    next_row = 2

    # ---- Top section: one row per Old, in Old's input order ----
    for old in olds:
        match = matches.get(old.sheet_index)
        if match is None:
            # No paired New row above the threshold → status missing.
            row_vals = [
                old.jama_id, old.item_type, old.name, old.description,
                "", "", "", "",
                ST_MISSING, "", "", "no New candidate above threshold",
            ]
            _write_row(ws, next_row, row_vals, wrap)
            _apply_hyperlink(ws, next_row, 1, old.hyperlink)
            counts[ST_MISSING] = counts.get(ST_MISSING, 0) + 1
            next_row += 1
            continue

        new_idx, name_sim = match
        new = news[new_idx]
        new_claimed.add(new_idx)
        d_sim = float(desc_sim[old.sheet_index, new_idx]) if desc_sim.size else 0.0

        v = verdicts.get((old.sheet_index, new_idx))
        if v is None:
            # We made the pairing but didn't classify (smoke test --limit, etc.).
            status = ST_ERROR
            reasoning = "no classification yet"
        else:
            status = v.status
            reasoning = v.reasoning

        row_vals = [
            old.jama_id, old.item_type, old.name, old.description,
            new.jama_id, new.item_type, new.name, new.description,
            status, round(name_sim, 4), round(d_sim, 4), reasoning,
        ]
        _write_row(ws, next_row, row_vals, wrap)
        _apply_hyperlink(ws, next_row, 1, old.hyperlink)
        _apply_hyperlink(ws, next_row, 5, new.hyperlink)
        counts[status] = counts.get(status, 0) + 1
        next_row += 1

    # ---- Bottom section: one row per unclaimed New, in New's input order ----
    for new in news:
        if new.sheet_index in new_claimed:
            continue
        row_vals = [
            "", "", "", "",
            new.jama_id, new.item_type, new.name, new.description,
            ST_UNMATCHED, "", "", "",
        ]
        _write_row(ws, next_row, row_vals, wrap)
        _apply_hyperlink(ws, next_row, 5, new.hyperlink)
        counts[ST_UNMATCHED] = counts.get(ST_UNMATCHED, 0) + 1
        next_row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(out_path)

    print("\nstatus counts:")
    for k in sorted(counts.keys()):
        print(f"  {k:<12} {counts[k]}")
    print(f"total rows in output: {next_row - 2}")


_HYPERLINK_FONT = Font(color="0563C1", underline="single")


def _write_row(ws, row_num: int, values: list, wrap: Alignment) -> None:
    for c_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=c_idx, value=val)
        cell.alignment = wrap
    status_cell = ws.cell(row=row_num, column=STATUS_COL_INDEX)
    fill = STATUS_FILLS.get(status_cell.value)
    if fill is not None:
        status_cell.fill = fill


def _apply_hyperlink(ws, row_num: int, col: int, url: str | None) -> None:
    if not url:
        return
    cell = ws.cell(row=row_num, column=col)
    cell.hyperlink = url
    cell.font = _HYPERLINK_FONT


# ---------- main ----------

def main() -> int:
    p = argparse.ArgumentParser(description="run4 — compare Old Jama vs New Jama names.")
    p.add_argument("xlsx", type=Path, help="input xlsx with Old Jama and New Jama sheets")
    p.add_argument("--out", type=Path, default=None,
                   help="output xlsx path (default: <input>_oldnew.xlsx)")
    p.add_argument("--old-sheet", default=OLD_SHEET_DEFAULT)
    p.add_argument("--new-sheet", default=NEW_SHEET_DEFAULT)
    p.add_argument("--threshold", type=float, default=DEFAULT_MATCH_THRESHOLD,
                   help="cosine-similarity floor below which an Old gets no candidate")
    p.add_argument("--re-extract", action="store_true")
    p.add_argument("--limit", type=int, default=None,
                   help="process only the first N pairs (smoke test)")
    p.add_argument("--concurrency", type=int, default=DEFAULT_CONCURRENCY)
    p.add_argument("--llm-url", default=DEFAULT_LLM_URL)
    p.add_argument("--llm-model", default=DEFAULT_LLM_MODEL)
    p.add_argument("--embed-model", default=DEFAULT_EMBED_MODEL)
    p.add_argument("--cls-checkpoint", type=Path, default=Path("classify.jsonl"))
    args = p.parse_args()

    if not args.xlsx.exists():
        print(f"ERR: xlsx not found: {args.xlsx}", file=sys.stderr)
        return 1

    work_dir = args.xlsx.parent
    out_path = args.out or (work_dir / f"{args.xlsx.stem}_oldnew.xlsx")

    old_json = work_dir / "old_jama.json"
    new_json = work_dir / "new_jama.json"

    olds = load_rows(args.xlsx, args.old_sheet, old_json, args.re_extract)
    news = load_rows(args.xlsx, args.new_sheet, new_json, args.re_extract)
    print(f"old: {len(olds)} rows | new: {len(news)} rows")
    if not olds or not news:
        print("nothing to do.", file=sys.stderr)
        return 1

    print(f"\n>>> embedding with {args.embed_model}...")
    embed_model = SentenceTransformer(args.embed_model, device="cpu")
    old_name_vecs = embed(embed_model, [r.name for r in olds], "old names")
    new_name_vecs = embed(embed_model, [r.name for r in news], "new names")
    sim = old_name_vecs @ new_name_vecs.T  # Name similarity (matches on this).

    # Description embeddings — used purely to populate the Description Similarity
    # column for matched pairs. Sparse Descriptions become tiny "" embeddings;
    # those rows will show low descriptional similarity, which is informative
    # rather than buggy.
    old_desc_vecs = embed(embed_model, [r.description or "" for r in olds], "old descriptions")
    new_desc_vecs = embed(embed_model, [r.description or "" for r in news], "new descriptions")
    desc_sim = old_desc_vecs @ new_desc_vecs.T

    matches = greedy_match(sim, args.threshold)
    print(f"matched pairs: {len(matches)} (threshold={args.threshold})")
    print(f"missing (Old with no candidate): {len(olds) - len(matches)}")
    print(f"will be unmatched (New not claimed): {len(news) - len(matches)}")

    # Build LLM jobs. Auto-identical now requires BOTH similarities to be
    # high — a row whose Name is unchanged but whose substantive Description
    # evolved should still get an LLM call so it can be flagged as a real
    # rewrite instead of silently passed.
    auto_identical: dict[tuple[int, int], Verdict] = {}
    jobs: list[tuple[Row, Row, float]] = []
    for old_idx, (new_idx, s) in matches.items():
        d = float(desc_sim[old_idx, new_idx])
        # If either side has an empty Description, fall back to name-only
        # auto-identical — there's nothing meaningful to verify on the
        # description axis.
        either_desc_empty = not olds[old_idx].description or not news[new_idx].description
        desc_ok = either_desc_empty or d >= AUTO_IDENTICAL_THRESHOLD
        if s >= AUTO_IDENTICAL_THRESHOLD and desc_ok:
            auto_identical[(old_idx, new_idx)] = Verdict(
                old_index=old_idx, new_index=new_idx,
                status=ST_IDENTICAL, similarity=float(s),
                reasoning=(
                    f"(auto-identical) name sim {s:.4f}, "
                    f"desc sim {d:.4f}{' (one side empty)' if either_desc_empty else ''}"
                ),
            )
        else:
            jobs.append((olds[old_idx], news[new_idx], float(s)))
    print(f"auto-identical (both name+desc sim ≥ {AUTO_IDENTICAL_THRESHOLD}): {len(auto_identical)}")
    print(f"LLM jobs: {len(jobs)}")

    if args.limit:
        jobs = jobs[:args.limit]
        print(f"--limit {args.limit}: processing first {len(jobs)} pairs")

    cls_ckpt_path = work_dir / args.cls_checkpoint
    cls_done_records = load_jsonl(cls_ckpt_path)
    cls_done_keys = {(d["old_index"], d["new_index"]) for d in cls_done_records
                     if d.get("status") != ST_ERROR}
    error_in_ckpt = sum(1 for d in cls_done_records if d.get("status") == ST_ERROR)
    print(f"classify checkpoint: {len(cls_done_records)} records "
          f"({error_in_ckpt} error rows will be retried)")

    http_client = httpx.AsyncClient(trust_env=False, timeout=httpx.Timeout(DEFAULT_LLM_TIMEOUT))
    client = AsyncOpenAI(base_url=args.llm_url, api_key="sk-local", http_client=http_client)

    print(f"\n>>> classifying with {args.llm_model}...")
    fresh_cls = asyncio.run(run_classify(
        client, args.llm_model, jobs, cls_ckpt_path, cls_done_keys, args.concurrency,
    ))
    cls_records = cls_done_records + [asdict(v) for v in fresh_cls]
    verdicts: dict[tuple[int, int], Verdict] = dict(auto_identical)
    for d in cls_records:
        loaded_status = str(d.get("status", ""))
        verdicts[(d["old_index"], d["new_index"])] = Verdict(
            old_index=d["old_index"],
            new_index=d["new_index"],
            status=loaded_status,
            similarity=float(d.get("similarity", 0.0)),
            reasoning=str(d.get("reasoning", "")),
        )

    write_report(out_path, olds, news, matches, verdicts, desc_sim)
    print(f"\nreport written -> {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
