import argparse
import json
import re
import sys
from pathlib import Path

import pandas as pd


HYPERLINK_FORMULA_RE = re.compile(r'HYPERLINK\(\s*"([^"]+)"', re.IGNORECASE)


def hyperlinks_xls(xlsx_path: Path, sheet_ref: str | int, header: int) -> dict[int, dict[str, str]]:
    """Extract hyperlinks from a .xls file via xlrd, keyed by data-row index."""
    import xlrd  # type: ignore

    book = xlrd.open_workbook(str(xlsx_path), formatting_info=True)
    sheet = book.sheet_by_name(sheet_ref) if isinstance(sheet_ref, str) else book.sheet_by_index(sheet_ref)
    col_names = [str(sheet.cell_value(header, c)) or f"col_{c}" for c in range(sheet.ncols)]

    result: dict[int, dict[str, str]] = {}
    for (r, c), hl in getattr(sheet, "hyperlink_map", {}).items():
        if r <= header:
            continue
        data_idx = r - header - 1
        url = getattr(hl, "url_or_path", None) or getattr(hl, "textmark", None)
        if not url:
            continue
        col = col_names[c] if c < len(col_names) else f"col_{c}"
        result.setdefault(data_idx, {})[col] = url
    return result


def hyperlinks_xlsx(xlsx_path: Path, sheet_ref: str | int, header: int) -> dict[int, dict[str, str]]:
    """Extract hyperlinks from a .xlsx file.

    Handles two storage forms Jama uses:
      - true hyperlinks (cell.hyperlink.target)
      - =HYPERLINK("url", "text") formulas (parsed from the formula string)
    """
    from openpyxl import load_workbook  # type: ignore

    # data_only=False so formula strings are returned as cell.value instead of the cached result.
    wb = load_workbook(xlsx_path, read_only=False, data_only=False)
    ws = wb[sheet_ref] if isinstance(sheet_ref, str) else wb.worksheets[sheet_ref]

    header_1i = header + 1
    try:
        header_row = next(ws.iter_rows(min_row=header_1i, max_row=header_1i))
    except StopIteration:
        return {}
    header_cells = list(header_row)
    col_names = [str(c.value) if c.value is not None else f"col_{c.column}" for c in header_cells]

    result: dict[int, dict[str, str]] = {}
    for r_idx, row in enumerate(ws.iter_rows(min_row=header_1i + 1)):
        for cell in row:
            url: str | None = None
            if cell.hyperlink and cell.hyperlink.target:
                url = cell.hyperlink.target
            else:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    m = HYPERLINK_FORMULA_RE.search(v)
                    if m:
                        url = m.group(1)
            if url:
                ci = cell.column - 1
                col = col_names[ci] if ci < len(col_names) else f"col_{cell.column}"
                result.setdefault(r_idx, {})[col] = url
    return result


def extract(xlsx_path: Path, sheet: str | int, out_path: Path, header: int) -> None:
    df = pd.read_excel(xlsx_path, sheet_name=sheet, dtype=str, header=header)
    df = df.dropna(axis=1, how="all")
    df = df.fillna("")
    rows = df.to_dict(orient="records")

    ext = xlsx_path.suffix.lower()
    try:
        if ext == ".xls":
            links = hyperlinks_xls(xlsx_path, sheet, header)
        elif ext == ".xlsx":
            links = hyperlinks_xlsx(xlsx_path, sheet, header)
        else:
            links = {}
            print(f"unknown extension {ext}, skipping hyperlink extraction", file=sys.stderr)
    except Exception as e:  # noqa: BLE001
        links = {}
        print(f"hyperlink extraction failed: {e}. Continuing without links.", file=sys.stderr)

    linked_row_count = 0
    for i, row in enumerate(rows):
        if i in links:
            row["_hyperlinks"] = links[i]
            linked_row_count += 1

    payload = {
        "source": str(xlsx_path),
        "sheet": sheet,
        "header_row": header + 1,
        "columns": list(df.columns),
        "row_count": len(rows),
        "rows_with_hyperlinks": linked_row_count,
        "rows": rows,
    }

    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2))
    print(f"wrote {len(rows)} rows, {len(df.columns)} columns -> {out_path}")
    print(f"header row: {header + 1}")
    print(f"rows with hyperlinks: {linked_row_count}")
    print(f"columns: {list(df.columns)}")
    if rows:
        print(f"first row sample: {dict(list(rows[0].items())[:5])}")
        sample_link = next((r["_hyperlinks"] for r in rows if "_hyperlinks" in r), None)
        if sample_link:
            print(f"first hyperlink sample: {sample_link}")


def main() -> int:
    p = argparse.ArgumentParser(description="Dump a Jama Excel export to JSON for inspection.")
    p.add_argument("xlsx", type=Path)
    p.add_argument("-o", "--out", type=Path, default=Path("excel.json"))
    p.add_argument("-s", "--sheet", default=0, help="sheet name or 0-based index (default: 0)")
    p.add_argument("--header", type=int, default=0,
                   help="0-indexed row containing column names. Jama often has metadata on the first few rows — "
                        "if your real headers are on row 4, pass --header 3 (default: 0)")
    args = p.parse_args()

    sheet: str | int
    try:
        sheet = int(args.sheet)
    except (TypeError, ValueError):
        sheet = args.sheet

    extract(args.xlsx, sheet, args.out, args.header)
    return 0


if __name__ == "__main__":
    sys.exit(main())
