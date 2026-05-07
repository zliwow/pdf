"""run3 — Name-only Jama vs PDF coverage check.

The customer asked for a focused comparison: take just the Jama Name field and
check whether the PDF substantively establishes/discusses what the Name asserts.
There is no Description column on this template, so embedding and the LLM see
only the Name.

Differences from run2:
  * Input has an Item Type column → filter Folder/Heading/Test directly.
  * Pre-classifier removed (no figure/table flag).
  * Embed Name only.
  * Classify prompt rewritten for substantive-coverage semantics — a passing
    topic mention is NOT identical; the PDF must support what the Name
    asserts.
  * Status is binary: identical | mismatch (mismatch covers both
    contradicting content and absence-from-PDF).
  * 64k max_tokens with /no_think fallback (sglang launched at 131k context).
  * Sheet selection prefers the customer's "Name vs PDF" tab.

Output writes into the input xlsx, columns D / E / F (PDF text /
Name vs PDF Status / difference). Every input row is preserved.

Usage:
  python run3.py input.xlsx input.pdf
  python run3.py input.xlsx input.pdf --limit 30
"""

import argparse
import asyncio
import difflib
import json
import re
import shutil
import sys
from dataclasses import asdict, dataclass
from pathlib import Path

import httpx
import numpy as np
from openai import AsyncOpenAI
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from sentence_transformers import SentenceTransformer
from tqdm import tqdm

import extract_excel
import extract_pdf


# ---------- config ----------

DEFAULT_LLM_URL = "http://localhost:8000/v1"
DEFAULT_LLM_MODEL = "Qwen/Qwen3.6-27B"
DEFAULT_EMBED_MODEL = "BAAI/bge-small-en-v1.5"
DEFAULT_TOP_K = 5
DEFAULT_CONCURRENCY = 16
DEFAULT_PAGE_TEXT_LIMIT = 8000
DEFAULT_LLM_TIMEOUT = 240.0  # bumped from run2's 180s — bigger token budget = longer calls.

# Sheet name candidates in priority order. Customer's tab is "Name vs PDF";
# fall back to first sheet if nothing matches.
PREFERRED_SHEET_NAMES = ["Name vs PDF", "Name vs Pdf", "Name vs PDF Status", "Sheet1"]

# Item types that are not real reqs — write the type verbatim into Status, no LLM.
SKIP_TYPES = {
    "Folder", "Heading", "Text",
    "Set", "Component",
    "Test Case", "Test Step",
}
VERIFICATION_SUBSTR = "verification"

# Column-header candidates (case-insensitive lookup).
ID_COL_CANDIDATES = ["ID", "Global ID", "GID", "Jama ID", "Item ID"]
NAME_COL_CANDIDATES = ["Name", "Title"]
TYPE_COL_CANDIDATES = ["Item Type", "Type", "Category"]

PDF_TEXT_COL_CANDIDATES = ["PDF text", "PDF Text", "PDF Text Content", "PDF Content"]
STATUS_COL_CANDIDATES = ["Name vs PDF Status", "Name Vs PDF Status", "Status"]
KEY_DIFF_COL_CANDIDATES = ["difference", "Difference", "Key Difference", "Key difference", "Diff"]

# Status vocabulary — binary like the second half of run2.
ST_IDENTICAL = "identical"
ST_MISMATCH = "mismatch"
ST_ERROR = "error"

STATUS_FILLS = {
    ST_IDENTICAL: PatternFill("solid", fgColor="C6EFCE"),  # green
    ST_MISMATCH:  PatternFill("solid", fgColor="FFEB9C"),  # yellow
    ST_ERROR:     PatternFill("solid", fgColor="F4B084"),  # orange
}
SKIP_TYPE_FILL = PatternFill("solid", fgColor="D9D9D9")    # grey for Folder/Heading/etc.

# Defensive ID-pattern fallback for folders, in case Item Type is ever blank.
FOLDER_ID_PATTERN = re.compile(r"-FLD-\d", re.IGNORECASE)


# ---------- data ----------

@dataclass
class Req:
    row_index: int        # 0-based data-row index (matches extract_excel.rows[i])
    sheet_row: int        # 1-based excel row number for writing back
    jama_id: str
    name: str
    item_type: str
    is_skip_type: bool
    skip_status: str      # value to write into Status when is_skip_type=True
    embed_text: str       # what we embed and feed the LLM (the Name only)


@dataclass
class Page:
    index: int
    page_number: int      # 1-based local page number within the slice
    text: str


@dataclass
class Verdict:
    row_index: int
    status: str           # identical / mismatch / error
    matched_page: int | None
    quoted_text: str
    reasoning: str


# ---------- helpers ----------

def is_skip_item_type(item_type: str) -> bool:
    t = (item_type or "").strip()
    if t in SKIP_TYPES:
        return True
    if VERIFICATION_SUBSTR in t.lower():
        return True
    return False


def looks_like_folder_id(jama_id: str) -> bool:
    return bool(FOLDER_ID_PATTERN.search(jama_id or ""))


def find_header_col(headers: list[str], candidates: list[str]) -> int:
    """Return 0-based index of the first header matching any candidate
    (case-insensitive). Falls back to substring match. Returns -1 if nothing fits."""
    norm = [(h.strip().lower() if h else "", i) for i, h in enumerate(headers)]
    for c in candidates:
        cn = c.strip().lower()
        for h, i in norm:
            if h == cn:
                return i
    for c in candidates:
        cn = c.strip().lower()
        for h, i in norm:
            if h and (cn in h or h in cn):
                return i
    return -1


def word_diff(a: str, b: str) -> str:
    """Compact word-level diff between Jama text and PDF quote."""
    if not a or not b:
        return ""
    aw = re.findall(r"\S+", a)
    bw = re.findall(r"\S+", b)
    sm = difflib.SequenceMatcher(a=aw, b=bw, autojunk=False)
    parts: list[str] = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            continue
        left = " ".join(aw[i1:i2])
        right = " ".join(bw[j1:j2])
        if tag == "replace":
            parts.append(f"[{left} -> {right}]")
        elif tag == "delete":
            parts.append(f"[-{left}]")
        elif tag == "insert":
            parts.append(f"[+{right}]")
    return " ".join(parts)


def detect_sheet(xlsx_path: Path) -> str | int:
    """Pick the customer's preferred sheet if present, else the first sheet."""
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        names = list(wb.sheetnames)
        wb.close()
    except Exception:  # noqa: BLE001
        return 0
    lowered = {n.strip().lower(): n for n in names}
    for preferred in PREFERRED_SHEET_NAMES:
        actual = lowered.get(preferred.strip().lower())
        if actual:
            return actual
    return 0


# ---------- extraction (cached) ----------

def ensure_extracted(xlsx_path: Path, pdf_path: Path, work_dir: Path,
                     sheet: str | int, re_extract: bool) -> tuple[Path, Path]:
    excel_json = work_dir / "excel.json"
    pdf_json = work_dir / "pdf.json"
    if re_extract or not excel_json.exists():
        print(f"extracting {xlsx_path.name} (sheet={sheet!r}) -> {excel_json.name} ...")
        extract_excel.extract(xlsx_path, sheet=sheet, out_path=excel_json, header=0)
    else:
        print(f"reusing existing {excel_json.name}")
    if re_extract or not pdf_json.exists():
        print(f"extracting {pdf_path.name} -> {pdf_json.name} ...")
        extract_pdf.extract(pdf_path, pdf_json, preview_chars=200,
                            start_page=None, end_page=None)
    else:
        print(f"reusing existing {pdf_json.name}")
    return excel_json, pdf_json


def load_reqs(excel_json: Path) -> tuple[list[Req], dict, list[str]]:
    data = json.loads(excel_json.read_text())
    rows = data["rows"]
    columns: list[str] = data["columns"]
    if not rows:
        return [], {}, columns

    sample = rows[0]

    def first_key(candidates: list[str]) -> str | None:
        for c in candidates:
            if c in sample:
                return c
        lower = {k.lower(): k for k in sample}
        for c in candidates:
            if c.lower() in lower:
                return lower[c.lower()]
        return None

    id_key = first_key(ID_COL_CANDIDATES) or "ID"
    name_key = first_key(NAME_COL_CANDIDATES) or "Name"
    type_key = first_key(TYPE_COL_CANDIDATES) or "Item Type"

    reqs: list[Req] = []
    for i, row in enumerate(rows):
        jama_id = str(row.get(id_key, "")).strip()
        name = str(row.get(name_key, "")).strip()
        itype = str(row.get(type_key, "")).strip()

        if is_skip_item_type(itype):
            skip = True
            skip_status = itype
        elif looks_like_folder_id(jama_id):
            # Defensive — shouldn't normally fire on this template since Item Type
            # is present, but keeps us robust if a Folder row ever loses its type.
            skip = True
            skip_status = "Folder"
        else:
            skip = False
            skip_status = ""

        # Name only — that's the explicit ask. ID is included in the LLM prompt
        # but not in the embedding (would be noise).
        embed_text = name

        reqs.append(Req(
            row_index=i,
            sheet_row=i + 2,  # header on row 1, data starts row 2
            jama_id=jama_id,
            name=name,
            item_type=itype,
            is_skip_type=skip,
            skip_status=skip_status,
            embed_text=embed_text,
        ))

    meta = {"id_key": id_key, "name_key": name_key, "type_key": type_key}
    return reqs, meta, columns


def load_pages(pdf_json: Path) -> list[Page]:
    data = json.loads(pdf_json.read_text())
    pages: list[Page] = []
    for s in data["sections"]:
        text = (s.get("text") or "").strip()
        if not text:
            continue
        pages.append(Page(
            index=s["index"],
            page_number=s["start_page"] + 1,
            text=text,
        ))
    return pages


# ---------- embedding ----------

def embed(model: SentenceTransformer, texts: list[str], label: str) -> np.ndarray:
    if not texts:
        return np.zeros((0, model.get_sentence_embedding_dimension()), dtype=np.float32)
    print(f"embedding {len(texts)} {label} on CPU...")
    vecs = model.encode(
        texts, batch_size=32, convert_to_numpy=True,
        normalize_embeddings=True, show_progress_bar=True,
    )
    return vecs.astype(np.float32)


def top_k_indices(sim_row: np.ndarray, k: int) -> list[int]:
    if k >= len(sim_row):
        return list(np.argsort(-sim_row))
    part = np.argpartition(-sim_row, k)[:k]
    return [int(i) for i in part[np.argsort(-sim_row[part])]]


# ---------- LLM prompt ----------

CLASSIFY_SYSTEM = (
    "You are a requirements traceability analyst. You will be given a Jama "
    "requirement Name and several candidate PDF pages retrieved by similarity. "
    "Decide whether the PDF substantively covers what the Name asserts.\n\n"
    "IMPORTANT — read carefully: a Jama Name is a short, declarative summary "
    "of a requirement, not just a topic label. Your job is NOT to check "
    "whether the PDF *mentions the topic* — it is to check whether the PDF "
    "*establishes the specific assertion the Name makes*.\n\n"
    "Examples:\n"
    "- Name 'BIST is performed' → identical only if a PDF page actually "
    "  establishes that BIST is performed (e.g. 'A built-in self test (BIST) "
    "  shall execute at startup'). A page that merely mentions BIST in passing "
    "  without establishing that it is performed is NOT identical.\n"
    "- Name 'Internal Clocks' → identical if a PDF page substantively "
    "  discusses internal clocks (a section on internal clocking, a paragraph "
    "  describing the internal oscillator). A passing word match is not "
    "  enough.\n"
    "- Name 'Clock Output and RTC Shall Use the Crystal Oscillator as Time "
    "  Base' → identical only if the PDF supports that specific assertion: "
    "  the clock output and RTC use the crystal oscillator as their time "
    "  base.\n\n"
    "Status definitions — BINARY (only two possible values):\n"
    "- identical: a candidate page substantively establishes/discusses what "
    "  the Name asserts. Different wording is FINE — what matters is whether "
    "  a reader looking up this requirement in the PDF would find supporting "
    "  content. Set matched_page to that page and quoted_text to the exact "
    "  substring from that page that supports the requirement.\n"
    "- mismatch: anything else. This covers BOTH:\n"
    "    (a) A candidate page discusses the same topic but contradicts the "
    "        Name (different signal, opposite behavior, conflicting "
    "        assertion). Set matched_page to that page and quoted_text to "
    "        the conflicting text.\n"
    "    (b) NO candidate page substantively covers what the Name asserts — "
    "        even if a page mentions the topic in passing, even if a related "
    "        concept appears nearby. Set matched_page=null and "
    "        quoted_text=\"\".\n\n"
    "When in doubt between identical and mismatch case (a), prefer "
    "identical. A wording difference alone is never a mismatch — only "
    "contradicting content is.\n\n"
    "When in doubt between identical and mismatch case (b), be STRICT: a "
    "passing topic mention is mismatch (b), not identical. The Name must be "
    "substantively supported by the PDF text.\n\n"
    "Output ONLY a single JSON object — no preamble, no code fences. Start "
    "with { and end with }."
)

CLASSIFY_USER = """Jama requirement:
ID: {jama_id}
Item Type: {item_type}
Name: {name}

Candidate PDF pages (each page text is enclosed in triple quotes):
{pages_block}

Return JSON:
{{
  "status": "identical" | "mismatch",
  "matched_page": <page number from above, or null if no candidate page substantively covers the Name>,
  "quoted_text": "<exact substring from the matched page that supports or contradicts the Name; empty string if no page covers it>",
  "reasoning": "<2-3 sentences explaining the classification>"
}}
"""


def parse_llm_json(raw: str) -> dict:
    s = raw.strip()
    if s.startswith("```"):
        s = re.sub(r"^```(?:json)?\s*", "", s)
        s = re.sub(r"\s*```$", "", s)
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", s, re.DOTALL)
        if not m:
            raise
        return json.loads(m.group(0))


def llm_content(msg) -> str:
    """With qwen3 reasoning parser the JSON lands in `content`; we keep
    `reasoning_content` as a fallback in case the server is launched without
    the parser flag."""
    c = (msg.content or "").strip()
    if c:
        return c
    return (getattr(msg, "reasoning_content", None) or "").strip()


# ---------- LLM call with /no_think fallback ----------

# Reasoning models occasionally burn the entire token budget on `<think>` and
# never emit the JSON. Two-attempt strategy: full reasoning first, then a
# /no_think retry that forces a direct answer (Qwen3 honors this token).
NO_THINK_TOKEN = "\n\n/no_think"


def _disable_thinking(messages: list[dict]) -> list[dict]:
    out = [dict(m) for m in messages]
    for m in reversed(out):
        if m.get("role") == "user":
            m["content"] = (m.get("content") or "") + NO_THINK_TOKEN
            break
    return out


async def _llm_call_with_fallback(
    client: AsyncOpenAI, model: str, messages: list[dict],
    max_tokens: int, fallback_max_tokens: int = 4000,
) -> tuple[str, dict]:
    diag: dict = {}

    # Attempt 1 — full reasoning.
    resp = await client.chat.completions.create(
        model=model, messages=messages,
        temperature=0.0, max_tokens=max_tokens,
    )
    choice = resp.choices[0]
    content = llm_content(choice.message)
    diag["a1_finish"] = choice.finish_reason
    diag["a1_reason_len"] = len(getattr(choice.message, "reasoning_content", "") or "")
    if content:
        return content, diag

    # Attempt 2 — /no_think fallback. Only reached if attempt 1 emitted empty content.
    # We additionally force JSON output via response_format={"type": "json_object"}.
    # Without this, the model has been observed to "stop" cleanly with both content
    # and reasoning_content empty — JSON mode rules out that no-op exit.
    fb_messages = _disable_thinking(messages)
    try:
        try:
            resp = await client.chat.completions.create(
                model=model, messages=fb_messages,
                temperature=0.0, max_tokens=fallback_max_tokens,
                response_format={"type": "json_object"},
            )
        except Exception as e_rf:  # noqa: BLE001
            # Some sglang builds may reject response_format. Retry without it so the
            # call still goes through; we lose the JSON guarantee but keep liveness.
            diag["a2_no_response_format"] = f"{type(e_rf).__name__}: {e_rf}"
            resp = await client.chat.completions.create(
                model=model, messages=fb_messages,
                temperature=0.0, max_tokens=fallback_max_tokens,
            )
        choice = resp.choices[0]
        content = llm_content(choice.message)
        diag["a2_finish"] = choice.finish_reason
        diag["a2_reason_len"] = len(getattr(choice.message, "reasoning_content", "") or "")
        diag["a2_no_think"] = True
        if content:
            return content, diag
    except Exception as e:  # noqa: BLE001
        diag["a2_error"] = f"{type(e).__name__}: {e}"

    return "", diag


# ---------- classify ----------

async def classify_one(client: AsyncOpenAI, model: str, req: Req,
                       candidates: list[Page], page_text_limit: int) -> Verdict:
    pages_block_parts = []
    for p in candidates:
        text = p.text[:page_text_limit]
        pages_block_parts.append(f"[Page {p.page_number}]\n\"\"\"\n{text}\n\"\"\"")
    pages_block = "\n\n".join(pages_block_parts)

    user = CLASSIFY_USER.format(
        jama_id=req.jama_id,
        item_type=req.item_type or "Customer Requirement",
        name=req.name[:1000],
        pages_block=pages_block,
    )
    messages = [
        {"role": "system", "content": CLASSIFY_SYSTEM},
        {"role": "user", "content": user},
    ]
    try:
        # 64k for the primary attempt — sglang's 131k context easily affords it.
        # /no_think fallback is bounded at 4k since there's no reasoning to grow.
        content, diag = await _llm_call_with_fallback(
            client, model, messages, max_tokens=64000, fallback_max_tokens=4000,
        )
    except Exception as e:  # noqa: BLE001
        return Verdict(row_index=req.row_index, status=ST_ERROR,
                       matched_page=None, quoted_text="",
                       reasoning=f"LLM call failed: {type(e).__name__}: {e}")
    if not content:
        return Verdict(row_index=req.row_index, status=ST_ERROR,
                       matched_page=None, quoted_text="",
                       reasoning=f"LLM empty after /no_think fallback. diag={diag}")
    try:
        data = parse_llm_json(content)
    except json.JSONDecodeError:
        return Verdict(row_index=req.row_index, status=ST_ERROR,
                       matched_page=None, quoted_text="",
                       reasoning=f"JSON parse failed. raw={content[:500]}")

    status = str(data.get("status", "")).strip().lower()
    if status == "missing":
        # Defensive: legacy vocabulary — coerce to mismatch.
        status = ST_MISMATCH
    if status not in {ST_IDENTICAL, ST_MISMATCH}:
        return Verdict(row_index=req.row_index, status=ST_ERROR,
                       matched_page=None, quoted_text="",
                       reasoning=f"unexpected status from LLM: {status!r}")

    matched_raw = data.get("matched_page")
    try:
        matched_page = int(matched_raw) if matched_raw not in (None, "", "null") else None
    except (TypeError, ValueError):
        matched_page = None

    return Verdict(
        row_index=req.row_index,
        status=status,
        matched_page=matched_page,
        quoted_text=str(data.get("quoted_text", ""))[:4000],
        reasoning=str(data.get("reasoning", ""))[:1000],
    )


# ---------- checkpoints ----------

def load_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        return []
    out = []
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            out.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return out


def append_jsonl(path: Path, obj: dict) -> None:
    with path.open("a") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")


# ---------- runner ----------

async def run_classify(client, model, jobs, checkpoint, done_keys,
                       concurrency, page_text_limit):
    sem = asyncio.Semaphore(concurrency)
    pbar = tqdm(total=len(jobs),
                initial=sum(1 for r, _ in jobs if r.row_index in done_keys),
                desc="classify", unit="req")
    results: list[Verdict] = []

    async def worker(req: Req, candidates: list[Page]):
        if req.row_index in done_keys:
            return
        async with sem:
            v = await classify_one(client, model, req, candidates, page_text_limit)
        append_jsonl(checkpoint, asdict(v))
        results.append(v)
        pbar.update(1)

    await asyncio.gather(*(worker(r, c) for r, c in jobs))
    pbar.close()
    return results


# ---------- writer ----------

def write_template(input_xlsx: Path, output_xlsx: Path, sheet_name: str | None,
                   reqs: list[Req], verdicts_by_row: dict[int, Verdict]) -> None:
    """Open input xlsx, fill PDF text / Status / difference per row, save as
    output_xlsx. Column A formulas (HYPERLINK) are preserved by openpyxl."""
    shutil.copyfile(input_xlsx, output_xlsx)
    wb = load_workbook(output_xlsx, data_only=False)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[0]

    headers_r1: list[str] = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers_r1.append(str(cell.value).strip() if cell.value is not None else "")

    # Some customer templates leave a row-1 header blank and put the column's
    # legend text on row 2 (e.g. "difference" sits below an empty header).
    # Use row 2 as a fallback when row 1 doesn't name a column we need.
    headers_r2: list[str] = []
    try:
        for cell in next(ws.iter_rows(min_row=2, max_row=2)):
            headers_r2.append(str(cell.value).strip() if cell.value is not None else "")
    except StopIteration:
        pass

    def locate(candidates: list[str]) -> int:
        idx = find_header_col(headers_r1, candidates)
        if idx >= 0:
            return idx
        return find_header_col(headers_r2, candidates)

    pdf_text_idx = locate(PDF_TEXT_COL_CANDIDATES)
    status_idx = locate(STATUS_COL_CANDIDATES)
    keydiff_idx = locate(KEY_DIFF_COL_CANDIDATES)

    missing = []
    if pdf_text_idx < 0:
        missing.append("PDF text")
    if status_idx < 0:
        missing.append("Name vs PDF Status")
    if keydiff_idx < 0:
        missing.append("difference")
    if missing:
        raise SystemExit(
            f"could not locate columns in input xlsx: {missing}. "
            f"row 1 headers: {headers_r1} | row 2 fallback: {headers_r2}"
        )

    pdf_text_col = pdf_text_idx + 1
    status_col = status_idx + 1
    keydiff_col = keydiff_idx + 1

    print(f"writing into sheet {ws.title!r}: PDF text={get_column_letter(pdf_text_col)}, "
          f"Status={get_column_letter(status_col)}, "
          f"difference={get_column_letter(keydiff_col)}")

    wrap = Alignment(wrap_text=True, vertical="top")
    counts: dict[str, int] = {}

    for req in reqs:
        if req.is_skip_type:
            label = req.skip_status or "Folder"
            sc = ws.cell(row=req.sheet_row, column=status_col, value=label)
            sc.fill = SKIP_TYPE_FILL
            sc.alignment = wrap
            counts[label] = counts.get(label, 0) + 1
            continue

        v = verdicts_by_row.get(req.row_index)
        if v is None:
            ws.cell(row=req.sheet_row, column=status_col, value="").alignment = wrap
            counts["unprocessed"] = counts.get("unprocessed", 0) + 1
            continue

        # PDF text cell — quote + (p.X) suffix when we have a matched page.
        pdf_text_value = ""
        if v.quoted_text and v.matched_page is not None:
            pdf_text_value = f"{v.quoted_text}\n\n(p.{v.matched_page})"
        elif v.quoted_text:
            pdf_text_value = v.quoted_text
        ws.cell(row=req.sheet_row, column=pdf_text_col, value=pdf_text_value).alignment = wrap

        # Status cell with color fill.
        sc = ws.cell(row=req.sheet_row, column=status_col, value=v.status)
        fill = STATUS_FILLS.get(v.status)
        if fill:
            sc.fill = fill
        sc.alignment = wrap

        # difference cell — Name-vs-quote diff is short by nature, so we
        # surface the LLM reasoning alongside it for context.
        if v.status == ST_IDENTICAL:
            diff = word_diff(req.name, v.quoted_text)
            kd = diff if diff else "(name supported by PDF)"
        elif v.status == ST_MISMATCH:
            diff = word_diff(req.name, v.quoted_text)
            kd = (diff + "\n\n" if diff else "") + v.reasoning
        else:
            kd = v.reasoning
        ws.cell(row=req.sheet_row, column=keydiff_col, value=kd[:1500]).alignment = wrap

        counts[v.status] = counts.get(v.status, 0) + 1

    wb.save(output_xlsx)

    print("\nstatus counts:")
    for k in sorted(counts.keys()):
        print(f"  {k:<20} {counts[k]}")


# ---------- main ----------

def main() -> int:
    p = argparse.ArgumentParser(description="run3 — Name-only Jama-vs-PDF coverage check.")
    p.add_argument("xlsx", type=Path, help="input Jama xlsx")
    p.add_argument("pdf", type=Path, help="input PDF")
    p.add_argument("--out", type=Path, default=None,
                   help="output xlsx path (default: <input>_filled.xlsx)")
    p.add_argument("--sheet", default=None,
                   help="explicit sheet name (default: 'Name vs PDF', then first sheet)")
    p.add_argument("--limit", type=int, default=None,
                   help="process only the first N requirement rows for a smoke test")
    p.add_argument("--re-extract", action="store_true",
                   help="force re-extracting excel.json / pdf.json from the inputs")
    p.add_argument("--top-k", type=int, default=DEFAULT_TOP_K)
    p.add_argument("--concurrency", type=int, default=DEFAULT_CONCURRENCY)
    p.add_argument("--llm-url", default=DEFAULT_LLM_URL)
    p.add_argument("--llm-model", default=DEFAULT_LLM_MODEL)
    p.add_argument("--embed-model", default=DEFAULT_EMBED_MODEL)
    p.add_argument("--page-text-limit", type=int, default=DEFAULT_PAGE_TEXT_LIMIT)
    p.add_argument("--cls-checkpoint", type=Path, default=Path("classify.jsonl"))
    args = p.parse_args()

    if not args.xlsx.exists():
        print(f"ERR: xlsx not found: {args.xlsx}", file=sys.stderr)
        return 1
    if not args.pdf.exists():
        print(f"ERR: pdf not found: {args.pdf}", file=sys.stderr)
        return 1

    work_dir = args.xlsx.parent
    out_path = args.out or (work_dir / f"{args.xlsx.stem}_filled.xlsx")

    sheet: str | int = args.sheet if args.sheet else detect_sheet(args.xlsx)
    print(f"using sheet: {sheet!r}")

    excel_json, pdf_json = ensure_extracted(args.xlsx, args.pdf, work_dir, sheet, args.re_extract)

    reqs, meta, columns = load_reqs(excel_json)
    pages = load_pages(pdf_json)
    print(f"loaded {len(reqs)} rows | {len(pages)} pages | column keys: {meta}")

    pre_filter_count = len(reqs)
    reqs = [r for r in reqs if r.jama_id]
    dropped_no_id = pre_filter_count - len(reqs)
    if dropped_no_id:
        print(f"dropped {dropped_no_id} rows with no Jama ID (left blank in output)")

    if args.limit:
        reqs = reqs[:args.limit]
        print(f"--limit {args.limit}: processing first {len(reqs)} rows")

    if not reqs or not pages:
        print("nothing to do.", file=sys.stderr)
        return 1

    real_reqs = [r for r in reqs if not r.is_skip_type]
    print(f"non-skip-type rows to classify: {len(real_reqs)}")

    cls_ckpt_path = work_dir / args.cls_checkpoint
    cls_done_records = load_jsonl(cls_ckpt_path)
    # Records whose status is "error" should be retried on the next run rather
    # than treated as completed work — the next attempt may benefit from prompt
    # changes or the json_object fallback.
    cls_done_keys = {d["row_index"] for d in cls_done_records
                     if d.get("status") != ST_ERROR}
    error_in_ckpt = sum(1 for d in cls_done_records if d.get("status") == ST_ERROR)
    print(f"classify checkpoint: {len(cls_done_records)} records "
          f"({error_in_ckpt} error rows will be retried)")

    # trust_env=False stops httpx from picking up corp HTTP_PROXY vars.
    http_client = httpx.AsyncClient(trust_env=False, timeout=httpx.Timeout(DEFAULT_LLM_TIMEOUT))
    client = AsyncOpenAI(base_url=args.llm_url, api_key="sk-local", http_client=http_client)

    sheet_name_for_writer: str | None = sheet if isinstance(sheet, str) else None

    if not real_reqs:
        print("no rows to classify; writing template now.")
        write_template(args.xlsx, out_path, sheet_name_for_writer, reqs, {})
        print(f"\nreport written -> {out_path}")
        return 0

    print(f"\n>>> embedding with {args.embed_model}...")
    embed_model = SentenceTransformer(args.embed_model, device="cpu")
    req_vecs = embed(embed_model, [r.embed_text for r in real_reqs], "reqs")
    page_vecs = embed(embed_model, [p.text for p in pages], "pages")
    sim = req_vecs @ page_vecs.T

    jobs: list[tuple[Req, list[Page]]] = []
    for i, req in enumerate(real_reqs):
        idxs = top_k_indices(sim[i], args.top_k)
        candidates = [pages[j] for j in idxs]
        jobs.append((req, candidates))
    print(f"classify jobs: {len(jobs)} (top-{args.top_k} pages each)")

    print(f"\n>>> classifying with {args.llm_model}...")
    fresh_cls = asyncio.run(run_classify(
        client, args.llm_model, jobs, cls_ckpt_path, cls_done_keys,
        args.concurrency, args.page_text_limit,
    ))
    cls_records = cls_done_records + [asdict(v) for v in fresh_cls]
    verdicts_by_row: dict[int, Verdict] = {}
    for d in cls_records:
        loaded_status = str(d.get("status", ""))
        if loaded_status == "missing":
            loaded_status = ST_MISMATCH
        verdicts_by_row[d["row_index"]] = Verdict(
            row_index=d["row_index"],
            status=loaded_status,
            matched_page=d.get("matched_page"),
            quoted_text=str(d.get("quoted_text", "")),
            reasoning=str(d.get("reasoning", "")),
        )

    write_template(args.xlsx, out_path, sheet_name_for_writer, reqs, verdicts_by_row)
    print(f"\nreport written -> {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
