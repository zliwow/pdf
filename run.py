"""Orchestrate the Excel vs PDF coverage check.

Pipeline:
  1. Load excel.json, pdf.json, toc.json.
  2. Filter Excel rows by Item Type (skip list).
  3. Embed filtered reqs + PDF pages on CPU with bge-small-en-v1.5.
  4. Retrieve top-K PDF pages per req by cosine similarity.
  5. Ask Qwen (sglang @ :8000, OpenAI-compatible) to classify each (req, page) pair.
  6. Aggregate best status per req, attach section label from ToC.
  7. Write report.xlsx with hyperlinked Jama IDs and colored status cells.

Resumable: every LLM response is appended to a JSONL checkpoint. Re-running skips
pairs already classified, so an interrupted run picks up where it left off.
"""

import argparse
import asyncio
import difflib
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path

import httpx
import numpy as np
from openai import AsyncOpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sentence_transformers import SentenceTransformer
from tqdm import tqdm


# ---------- config defaults ----------

DEFAULT_SKIP_TYPES = {
    "Folder", "Text", "Heading",
    "Test Case", "Test Step",
    "Set", "Component",
}
VERIFICATION_SUBSTR = "verification"  # any item type containing this is skipped

DEFAULT_EMBED_MODEL = "BAAI/bge-small-en-v1.5"
DEFAULT_LLM_MODEL = "Qwen/Qwen3.5-27B-FP8"
DEFAULT_LLM_URL = "http://localhost:8000/v1"
DEFAULT_TOP_K = 5
DEFAULT_CONCURRENCY = 8

ID_COL_CANDIDATES = ["ID", "Global ID", "GID", "Jama ID", "Item ID"]
NAME_COL_CANDIDATES = ["Name", "Title"]
DESC_COL_CANDIDATES = ["Description", "Text", "Requirement"]
TYPE_COL_CANDIDATES = ["Item Type", "Type", "Category"]


# ---------- data classes ----------

@dataclass
class Req:
    row_index: int
    jama_id: str
    name: str
    description: str
    item_type: str
    hyperlink: str | None
    embed_text: str  # what we embed / send to LLM


@dataclass
class Page:
    index: int              # matches pdf.json "sections" index
    page_number: int        # 1-indexed human page
    text: str


@dataclass
class Verdict:
    req_row: int
    page_index: int
    status: str             # covered | mismatch | not_mentioned | error
    quoted_text: str
    reasoning: str


# ---------- helpers ----------

def first_present(row: dict, candidates: list[str]) -> str | None:
    for c in candidates:
        if c in row:
            return c
    # case-insensitive fallback
    lower = {k.lower(): k for k in row}
    for c in candidates:
        if c.lower() in lower:
            return lower[c.lower()]
    return None


def should_skip(item_type: str, skip_types: set[str]) -> bool:
    t = (item_type or "").strip()
    if t in skip_types:
        return True
    if VERIFICATION_SUBSTR in t.lower():
        return True
    return False


def load_reqs(excel_path: Path, skip_types: set[str]) -> tuple[list[Req], dict]:
    data = json.loads(excel_path.read_text())
    rows = data["rows"]
    if not rows:
        return [], data

    sample = rows[0]
    id_col = first_present(sample, ID_COL_CANDIDATES) or "ID"
    name_col = first_present(sample, NAME_COL_CANDIDATES) or "Name"
    desc_col = first_present(sample, DESC_COL_CANDIDATES) or "Description"
    type_col = first_present(sample, TYPE_COL_CANDIDATES) or "Item Type"

    reqs: list[Req] = []
    for i, row in enumerate(rows):
        itype = str(row.get(type_col, "")).strip()
        if should_skip(itype, skip_types):
            continue
        jama_id = str(row.get(id_col, "")).strip()
        name = str(row.get(name_col, "")).strip()
        desc = str(row.get(desc_col, "")).strip()
        if not (jama_id or name or desc):
            continue
        links = row.get("_hyperlinks") or {}
        hyperlink = links.get(id_col) or next(iter(links.values()), None)
        embed_text = f"[{jama_id}] {name}\n{desc}".strip()
        reqs.append(Req(
            row_index=i,
            jama_id=jama_id,
            name=name,
            description=desc,
            item_type=itype,
            hyperlink=hyperlink,
            embed_text=embed_text,
        ))

    meta = {
        "id_col": id_col, "name_col": name_col,
        "desc_col": desc_col, "type_col": type_col,
    }
    return reqs, meta


def load_pages(pdf_path: Path) -> list[Page]:
    data = json.loads(pdf_path.read_text())
    pages: list[Page] = []
    for s in data["sections"]:
        text = (s.get("text") or "").strip()
        if not text:
            continue
        pages.append(Page(
            index=s["index"],
            page_number=s["start_page"] + 1,
            text=text,
        ))
    return pages


def load_toc(toc_path: Path | None) -> list[dict]:
    if toc_path is None or not toc_path.exists():
        return []
    data = json.loads(toc_path.read_text())
    return data.get("entries", [])


def section_for_page(page_number: int, toc: list[dict]) -> str:
    """Return a '§3.2.1 Boot Sequence' label for the deepest section containing the page."""
    best = None
    for e in toc:
        p = e.get("page")
        if p is None or p > page_number:
            continue
        if best is None or p > best.get("page", 0) or (
            p == best.get("page") and e.get("level", 0) > best.get("level", 0)
        ):
            best = e
    if not best:
        return ""
    return f"§{best['section_number']} {best['title']}"


def word_diff(a: str, b: str) -> str:
    """Compact word-level diff between excel text and PDF quote."""
    if not a or not b:
        return ""
    aw = re.findall(r"\S+", a)
    bw = re.findall(r"\S+", b)
    sm = difflib.SequenceMatcher(a=aw, b=bw, autojunk=False)
    parts: list[str] = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            continue
        left = " ".join(aw[i1:i2])
        right = " ".join(bw[j1:j2])
        if tag == "replace":
            parts.append(f"[{left} -> {right}]")
        elif tag == "delete":
            parts.append(f"[-{left}]")
        elif tag == "insert":
            parts.append(f"[+{right}]")
    return " ".join(parts)


# ---------- embedding + retrieval ----------

def embed(model: SentenceTransformer, texts: list[str], label: str) -> np.ndarray:
    if not texts:
        return np.zeros((0, model.get_sentence_embedding_dimension()), dtype=np.float32)
    print(f"embedding {len(texts)} {label} on CPU...")
    vecs = model.encode(
        texts, batch_size=32, convert_to_numpy=True,
        normalize_embeddings=True, show_progress_bar=True,
    )
    return vecs.astype(np.float32)


def top_k_indices(sim_row: np.ndarray, k: int) -> list[int]:
    if k >= len(sim_row):
        return list(np.argsort(-sim_row))
    part = np.argpartition(-sim_row, k)[:k]
    return [int(i) for i in part[np.argsort(-sim_row[part])]]


# ---------- LLM ----------

SYSTEM_PROMPT = (
    "You are a requirements traceability analyst. For each (requirement, PDF page) pair, "
    "decide whether the PDF page covers the requirement. Be strict about the text — if the "
    "page discusses the same topic but with different wording, values, or signal names, mark it "
    "as mismatch, not covered. Reply with a single JSON object and nothing else."
)

USER_TEMPLATE = """Requirement:
ID: {jama_id}
Name: {name}
Description: {description}

PDF page {page_number} text:
\"\"\"
{page_text}
\"\"\"

Return JSON with:
- status: one of "covered", "mismatch", "not_mentioned"
- quoted_text: exact substring of the PDF page text that matches this requirement, or "" if not_mentioned
- reasoning: one sentence explaining the classification
"""


def parse_llm_json(raw: str) -> dict:
    """Tolerant JSON parse — sglang usually returns clean JSON, but be safe."""
    s = raw.strip()
    if s.startswith("```"):
        s = re.sub(r"^```(?:json)?\s*", "", s)
        s = re.sub(r"\s*```$", "", s)
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", s, re.DOTALL)
        if not m:
            raise
        return json.loads(m.group(0))


async def classify_one(
    client: AsyncOpenAI, model: str, req: Req, page: Page,
    page_text_limit: int,
) -> Verdict:
    page_text = page.text[:page_text_limit]
    user = USER_TEMPLATE.format(
        jama_id=req.jama_id, name=req.name,
        description=req.description[:4000],
        page_number=page.page_number, page_text=page_text,
    )
    try:
        resp = await client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user},
            ],
            response_format={"type": "json_object"},
            temperature=0.0,
            max_tokens=400,
        )
        msg = resp.choices[0].message
        # Reasoning models sometimes put the payload in reasoning_content instead.
        content = msg.content or getattr(msg, "reasoning_content", None) or "{}"
        data = parse_llm_json(content)
        status = str(data.get("status", "")).strip().lower()
        if status not in {"covered", "mismatch", "not_mentioned"}:
            status = "error"
        return Verdict(
            req_row=req.row_index, page_index=page.index, status=status,
            quoted_text=str(data.get("quoted_text", ""))[:4000],
            reasoning=str(data.get("reasoning", ""))[:1000],
        )
    except Exception as e:  # noqa: BLE001
        return Verdict(
            req_row=req.row_index, page_index=page.index, status="error",
            quoted_text="", reasoning=f"LLM call failed: {type(e).__name__}: {e}",
        )


# ---------- checkpoint ----------

def load_checkpoint(path: Path) -> dict[tuple[int, int], Verdict]:
    if not path.exists():
        return {}
    done: dict[tuple[int, int], Verdict] = {}
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            d = json.loads(line)
            v = Verdict(**d)
        except Exception:  # noqa: BLE001
            continue
        done[(v.req_row, v.page_index)] = v
    return done


def append_checkpoint(path: Path, v: Verdict) -> None:
    with path.open("a") as f:
        f.write(json.dumps(v.__dict__, ensure_ascii=False) + "\n")


# ---------- aggregation ----------

STATUS_PRIORITY = {"covered": 3, "mismatch": 2, "not_mentioned": 1, "error": 0}


def aggregate(req: Req, verdicts: list[Verdict], pages_by_idx: dict[int, Page], toc: list[dict]) -> dict:
    if not verdicts:
        return {
            "req": req, "status": "not_mentioned", "matched_page": None,
            "matched_section": "", "pdf_quote": "", "reasoning": "No candidate pages.",
            "diff": "",
        }
    best = max(verdicts, key=lambda v: STATUS_PRIORITY.get(v.status, -1))
    # Only surface matched_page / matched_section / quote when the LLM actually matched
    # something. For not_mentioned and error, the top-K candidate is just a retrieval
    # artifact, not a real match — showing it is misleading.
    is_match = best.status in {"covered", "mismatch"}
    matched_page = pages_by_idx.get(best.page_index) if is_match else None
    page_num = matched_page.page_number if matched_page else None
    section = section_for_page(page_num, toc) if page_num else ""
    diff = word_diff(req.description, best.quoted_text) if is_match else ""
    return {
        "req": req,
        "status": best.status,
        "matched_page": page_num,
        "matched_section": section,
        "pdf_quote": best.quoted_text if is_match else "",
        "reasoning": best.reasoning,
        "diff": diff,
    }


# ---------- report ----------

STATUS_FILLS = {
    "covered":      PatternFill("solid", fgColor="C6EFCE"),  # green
    "mismatch":     PatternFill("solid", fgColor="FFEB9C"),  # yellow
    "not_mentioned":PatternFill("solid", fgColor="FFC7CE"),  # red
    "error":        PatternFill("solid", fgColor="D9D9D9"),  # grey
}

REPORT_COLUMNS = [
    ("Jama ID", 18),
    ("Name", 40),
    ("Item Type", 22),
    ("Status", 16),
    ("Matched Page", 14),
    ("Matched Section", 40),
    ("Excel Description", 60),
    ("PDF Quote", 60),
    ("Word Diff", 40),
    ("Reasoning", 60),
]


def write_report(results: list[dict], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Coverage"

    for col_idx, (name, width) in enumerate(REPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wrap = Alignment(wrap_text=True, vertical="top")

    for r_idx, res in enumerate(results, start=2):
        req: Req = res["req"]
        values = [
            req.jama_id, req.name, req.item_type,
            res["status"], res["matched_page"], res["matched_section"],
            req.description, res["pdf_quote"], res["diff"], res["reasoning"],
        ]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = wrap

        id_cell = ws.cell(row=r_idx, column=1)
        if req.hyperlink:
            id_cell.hyperlink = req.hyperlink
            id_cell.font = Font(color="0563C1", underline="single")

        status_cell = ws.cell(row=r_idx, column=4)
        fill = STATUS_FILLS.get(res["status"])
        if fill is not None:
            status_cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(out_path)


# ---------- main ----------

async def run_llm(
    client: AsyncOpenAI, model: str, pairs: list[tuple[Req, Page]],
    concurrency: int, checkpoint_path: Path, done: dict[tuple[int, int], Verdict],
    page_text_limit: int,
) -> list[Verdict]:
    sem = asyncio.Semaphore(concurrency)
    verdicts: list[Verdict] = list(done.values())
    pbar = tqdm(total=len(pairs), initial=sum(1 for r, p in pairs if (r.row_index, p.index) in done),
                desc="Qwen verify", unit="call")

    async def worker(req: Req, page: Page) -> None:
        key = (req.row_index, page.index)
        if key in done:
            return
        async with sem:
            v = await classify_one(client, model, req, page, page_text_limit)
        append_checkpoint(checkpoint_path, v)
        verdicts.append(v)
        pbar.update(1)

    await asyncio.gather(*(worker(r, p) for r, p in pairs))
    pbar.close()
    return verdicts


def build_pairs(reqs: list[Req], pages: list[Page], req_vecs: np.ndarray,
                page_vecs: np.ndarray, top_k: int) -> list[tuple[Req, Page]]:
    sim = req_vecs @ page_vecs.T  # (R, P), both normalized
    pages_by_idx_pos = {i: p for i, p in enumerate(pages)}
    pairs: list[tuple[Req, Page]] = []
    for i, req in enumerate(reqs):
        idxs = top_k_indices(sim[i], top_k)
        for j in idxs:
            pairs.append((req, pages_by_idx_pos[j]))
    return pairs


def parse_types(s: str | None) -> set[str]:
    if not s:
        return set(DEFAULT_SKIP_TYPES)
    return {t.strip() for t in s.split(",") if t.strip()}


def main() -> int:
    p = argparse.ArgumentParser(description="Run Excel vs PDF coverage check.")
    p.add_argument("--excel", type=Path, default=Path("excel.json"))
    p.add_argument("--pdf", type=Path, default=Path("pdf.json"))
    p.add_argument("--toc", type=Path, default=Path("toc.json"))
    p.add_argument("--out", type=Path, default=Path("report.xlsx"))
    p.add_argument("--checkpoint", type=Path, default=Path("results.jsonl"))
    p.add_argument("--skip-types", default=None,
                   help="comma-separated Item Types to skip (default: folder/text/test/etc.)")
    p.add_argument("--top-k", type=int, default=DEFAULT_TOP_K)
    p.add_argument("--concurrency", type=int, default=DEFAULT_CONCURRENCY)
    p.add_argument("--llm-url", default=DEFAULT_LLM_URL)
    p.add_argument("--llm-model", default=DEFAULT_LLM_MODEL)
    p.add_argument("--embed-model", default=DEFAULT_EMBED_MODEL)
    p.add_argument("--page-text-limit", type=int, default=6000,
                   help="truncate each page's text to this many chars before sending to LLM")
    p.add_argument("--limit", type=int, default=None,
                   help="process only the first N reqs (for smoke tests)")
    args = p.parse_args()

    skip_types = parse_types(args.skip_types)

    print("loading excel + pdf + toc...")
    reqs, meta = load_reqs(args.excel, skip_types)
    pages = load_pages(args.pdf)
    toc = load_toc(args.toc)
    print(f"reqs after filter: {len(reqs)} | pages: {len(pages)} | toc entries: {len(toc)}")
    print(f"columns used: {meta}")
    if args.limit:
        reqs = reqs[:args.limit]
        print(f"limit applied: processing {len(reqs)} reqs")

    if not reqs or not pages:
        print("nothing to do.", file=sys.stderr)
        return 1

    print(f"loading embedding model: {args.embed_model}")
    embed_model = SentenceTransformer(args.embed_model, device="cpu")
    req_vecs = embed(embed_model, [r.embed_text for r in reqs], "reqs")
    page_vecs = embed(embed_model, [p.text for p in pages], "pages")

    pairs = build_pairs(reqs, pages, req_vecs, page_vecs, args.top_k)
    print(f"pairs to classify: {len(pairs)} (reqs × top-{args.top_k})")

    done = load_checkpoint(args.checkpoint)
    if done:
        print(f"resuming from checkpoint: {len(done)} pairs already classified")

    # trust_env=False stops httpx from picking up HTTP_PROXY / HTTPS_PROXY vars,
    # which on corporate machines otherwise hijack localhost calls and break sglang.
    http_client = httpx.AsyncClient(trust_env=False, timeout=httpx.Timeout(120.0))
    client = AsyncOpenAI(base_url=args.llm_url, api_key="sk-local", http_client=http_client)
    verdicts = asyncio.run(run_llm(
        client, args.llm_model, pairs, args.concurrency,
        args.checkpoint, done, args.page_text_limit,
    ))

    # Index verdicts by req row for aggregation.
    by_req: dict[int, list[Verdict]] = {}
    for v in verdicts:
        by_req.setdefault(v.req_row, []).append(v)
    pages_by_idx = {p.index: p for p in pages}

    results = [aggregate(r, by_req.get(r.row_index, []), pages_by_idx, toc) for r in reqs]

    # Sort: errors first (so you see them), then not_mentioned, mismatch, covered.
    order = {"error": 0, "not_mentioned": 1, "mismatch": 2, "covered": 3}
    results.sort(key=lambda x: (order.get(x["status"], 99), x["req"].jama_id))

    write_report(results, args.out)

    counts: dict[str, int] = {}
    for r in results:
        counts[r["status"]] = counts.get(r["status"], 0) + 1
    print("\nsummary:")
    for status in ["covered", "mismatch", "not_mentioned", "error"]:
        print(f"  {status:<15} {counts.get(status, 0)}")
    print(f"\nreport written -> {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
