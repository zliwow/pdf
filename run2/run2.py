"""run2 — Excel vs PDF coverage check, fills the customer's template in place.

Differences from the previous run:
  * Output is the *input* xlsx with three columns filled (PDF Text Content / Status /
    Key Difference). Every input row is preserved — Folder/Heading/Test rows get the
    Item Type written verbatim into Status with a grey fill.
  * A pre-classifier LLM call per non-skip row decides whether the item is text-based
    or a figure/table-only reference. Figure/table rows are marked but not analyzed.
  * The page classifier batches all top-K candidate pages into a single LLM call per
    req — one round trip instead of K — and outputs the customer's vocabulary directly
    (identical / mismatch / missing).
  * Page numbers are local to the slice (1..N), with `(p.X)` appended to the PDF Text
    Content cell for quick navigation.

Usage:
  python run2.py input.xlsx input.pdf
  python run2.py input.xlsx input.pdf --limit 30        # smoke test
  python run2.py input.xlsx input.pdf --resume          # pick up from checkpoint

Resumable: every LLM response is appended to a JSONL file. Re-running with --resume
skips work already done.
"""

import argparse
import asyncio
import difflib
import json
import re
import shutil
import sys
from dataclasses import asdict, dataclass
from pathlib import Path

import httpx
import numpy as np
from openai import AsyncOpenAI
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sentence_transformers import SentenceTransformer
from tqdm import tqdm

import extract_excel
import extract_pdf


# ---------- config ----------

DEFAULT_LLM_URL = "http://localhost:8000/v1"
DEFAULT_LLM_MODEL = "Qwen/Qwen3.6-27B"
DEFAULT_EMBED_MODEL = "BAAI/bge-small-en-v1.5"
DEFAULT_TOP_K = 5
DEFAULT_CONCURRENCY = 16
DEFAULT_PAGE_TEXT_LIMIT = 8000  # per-page char cap before sending to LLM
DEFAULT_LLM_TIMEOUT = 180.0

# Item types that are not real reqs — write the type verbatim into Status, no LLM.
SKIP_TYPES = {
    "Folder", "Heading", "Text",
    "Set", "Component",
    "Test Case", "Test Step",
}
VERIFICATION_SUBSTR = "verification"  # any item type containing this is also skipped

# Column-header candidates (case-insensitive lookup).
ID_COL_CANDIDATES = ["ID", "Global ID", "GID", "Jama ID", "Item ID"]
NAME_COL_CANDIDATES = ["Name", "Title"]
DESC_COL_CANDIDATES = ["Description", "Text", "Requirement"]
TYPE_COL_CANDIDATES = ["Item Type", "Type", "Category"]
PDF_TEXT_COL_CANDIDATES = ["PDF Text Content", "PDF Text", "PDF Content"]
STATUS_COL_CANDIDATES = ["Status"]
KEY_DIFF_COL_CANDIDATES = ["Key Difference", "Key difference", "Difference", "Diff"]

# Status vocabulary written into the Status column. Customer asked for a binary
# decision: either the PDF faithfully covers the requirement (identical) or it
# doesn't (mismatch). The "doesn't cover" case includes both contradicting
# content and absence-from-PDF — the writer differentiates the two via the
# matched_page / quoted_text fields, but the Status cell only shows one of two
# values for text reqs.
ST_IDENTICAL = "identical"
ST_MISMATCH = "mismatch"
ST_FIGURE = "figure/table"
ST_ERROR = "error"

STATUS_FILLS = {
    ST_IDENTICAL: PatternFill("solid", fgColor="C6EFCE"),  # green
    ST_MISMATCH:  PatternFill("solid", fgColor="FFEB9C"),  # yellow
    ST_FIGURE:    PatternFill("solid", fgColor="DCE6F1"),  # light blue
    ST_ERROR:     PatternFill("solid", fgColor="F4B084"),  # orange
}
SKIP_TYPE_FILL = PatternFill("solid", fgColor="D9D9D9")    # grey for Folder/Heading/etc.


# ---------- data ----------

@dataclass
class Req:
    row_index: int        # 0-based data-row index (matches extract_excel.rows[i])
    sheet_row: int        # 1-based excel row number for writing back
    jama_id: str
    name: str
    description: str
    item_type: str
    is_skip_type: bool    # Folder/Heading/etc — bypass LLM
    skip_status: str      # what to write into the Status cell when is_skip_type=True
    embed_text: str       # what we embed and send to LLM


@dataclass
class Page:
    index: int            # matches pdf.json sections[i]['index']
    page_number: int      # 1-based local page number within the slice
    text: str


@dataclass
class PreVerdict:
    row_index: int
    is_text: bool         # False = figure/table-only reference
    reason: str
    raw: str = ""         # for debugging on parse failures


@dataclass
class Verdict:
    row_index: int
    status: str           # identical / mismatch / missing / error
    matched_page: int | None
    quoted_text: str
    reasoning: str


# ---------- helpers ----------

# Jama folder IDs follow the pattern <project>-FLD-<num>. We use this as a fallback
# when the input xlsx has no Item Type column (the customer's template does not
# carry one), so folders still get written out as `Folder` instead of going through
# the LLM. Match is case-insensitive on the embedded `-FLD-` token.
FOLDER_ID_PATTERN = re.compile(r"-FLD-\d", re.IGNORECASE)


def is_skip_item_type(item_type: str) -> bool:
    t = (item_type or "").strip()
    if t in SKIP_TYPES:
        return True
    if VERIFICATION_SUBSTR in t.lower():
        return True
    return False


def looks_like_folder_id(jama_id: str) -> bool:
    return bool(FOLDER_ID_PATTERN.search(jama_id or ""))


def find_header_col(headers: list[str], candidates: list[str]) -> int:
    """Return 0-based index of the first header matching any candidate (case-insensitive).
    Falls back to substring match. Returns -1 if nothing fits."""
    norm = [(h.strip().lower() if h else "", i) for i, h in enumerate(headers)]
    for c in candidates:
        cn = c.strip().lower()
        for h, i in norm:
            if h == cn:
                return i
    for c in candidates:
        cn = c.strip().lower()
        for h, i in norm:
            if h and (cn in h or h in cn):
                return i
    return -1


def word_diff(a: str, b: str) -> str:
    """Compact word-level diff between Jama text and PDF quote."""
    if not a or not b:
        return ""
    aw = re.findall(r"\S+", a)
    bw = re.findall(r"\S+", b)
    sm = difflib.SequenceMatcher(a=aw, b=bw, autojunk=False)
    parts: list[str] = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            continue
        left = " ".join(aw[i1:i2])
        right = " ".join(bw[j1:j2])
        if tag == "replace":
            parts.append(f"[{left} -> {right}]")
        elif tag == "delete":
            parts.append(f"[-{left}]")
        elif tag == "insert":
            parts.append(f"[+{right}]")
    return " ".join(parts)


# ---------- extraction (cached) ----------

def ensure_extracted(xlsx_path: Path, pdf_path: Path, work_dir: Path,
                     re_extract: bool) -> tuple[Path, Path]:
    excel_json = work_dir / "excel.json"
    pdf_json = work_dir / "pdf.json"
    if re_extract or not excel_json.exists():
        print(f"extracting {xlsx_path.name} -> {excel_json.name} ...")
        extract_excel.extract(xlsx_path, sheet=0, out_path=excel_json, header=0)
    else:
        print(f"reusing existing {excel_json.name}")
    if re_extract or not pdf_json.exists():
        print(f"extracting {pdf_path.name} -> {pdf_json.name} ...")
        extract_pdf.extract(pdf_path, pdf_json, preview_chars=200,
                            start_page=None, end_page=None)
    else:
        print(f"reusing existing {pdf_json.name}")
    return excel_json, pdf_json


def load_reqs(excel_json: Path) -> tuple[list[Req], dict, list[str]]:
    data = json.loads(excel_json.read_text())
    rows = data["rows"]
    columns: list[str] = data["columns"]
    if not rows:
        return [], {}, columns

    sample = rows[0]
    # Resolve which keys in `rows` correspond to ID / Name / Description / Item Type.
    def first_key(candidates: list[str]) -> str | None:
        for c in candidates:
            if c in sample:
                return c
        lower = {k.lower(): k for k in sample}
        for c in candidates:
            if c.lower() in lower:
                return lower[c.lower()]
        return None

    id_key = first_key(ID_COL_CANDIDATES) or "ID"
    name_key = first_key(NAME_COL_CANDIDATES) or "Name"
    desc_key = first_key(DESC_COL_CANDIDATES) or "Description"
    type_key = first_key(TYPE_COL_CANDIDATES) or "Item Type"

    reqs: list[Req] = []
    for i, row in enumerate(rows):
        jama_id = str(row.get(id_key, "")).strip()
        name = str(row.get(name_key, "")).strip()
        desc = str(row.get(desc_key, "")).strip()
        itype = str(row.get(type_key, "")).strip()
        # Two ways to qualify as skip: real Item Type, or ID-pattern fallback.
        # When Item Type is present we use it; otherwise we infer "Folder" from
        # IDs of the form *-FLD-<num>.
        if is_skip_item_type(itype):
            skip = True
            skip_status = itype
        elif looks_like_folder_id(jama_id):
            skip = True
            skip_status = "Folder"
        else:
            skip = False
            skip_status = ""
        embed_text = f"[{jama_id}] {name}\n{desc}".strip()
        reqs.append(Req(
            row_index=i,
            sheet_row=i + 2,  # header on row 1, data starts row 2
            jama_id=jama_id,
            name=name,
            description=desc,
            item_type=itype,
            is_skip_type=skip,
            skip_status=skip_status,
            embed_text=embed_text,
        ))

    meta = {"id_key": id_key, "name_key": name_key,
            "desc_key": desc_key, "type_key": type_key}
    return reqs, meta, columns


def load_pages(pdf_json: Path) -> list[Page]:
    data = json.loads(pdf_json.read_text())
    pages: list[Page] = []
    for s in data["sections"]:
        text = (s.get("text") or "").strip()
        if not text:
            continue
        pages.append(Page(
            index=s["index"],
            page_number=s["start_page"] + 1,
            text=text,
        ))
    return pages


# ---------- embedding ----------

def embed(model: SentenceTransformer, texts: list[str], label: str) -> np.ndarray:
    if not texts:
        return np.zeros((0, model.get_sentence_embedding_dimension()), dtype=np.float32)
    print(f"embedding {len(texts)} {label} on CPU...")
    vecs = model.encode(
        texts, batch_size=32, convert_to_numpy=True,
        normalize_embeddings=True, show_progress_bar=True,
    )
    return vecs.astype(np.float32)


def top_k_indices(sim_row: np.ndarray, k: int) -> list[int]:
    if k >= len(sim_row):
        return list(np.argsort(-sim_row))
    part = np.argpartition(-sim_row, k)[:k]
    return [int(i) for i in part[np.argsort(-sim_row[part])]]


# ---------- LLM prompts ----------

PRECLASS_SYSTEM = (
    "You are reviewing Jama requirement items. For each item, decide whether it can "
    "be text-compared against PDF body text, or whether its substantive content lives "
    "in an attachment (figure, table, diagram, schematic, image) that is NOT present "
    "in the description.\n\n"
    "Return type=\"figure_table\" ONLY when one of these is true:\n"
    "  1. The description is empty or near-empty AND the name refers to a visual "
    "     artifact, e.g. 'Block Diagram', 'Timing Diagram', 'State Diagram', "
    "     'Schematic', 'Figure 3.2', 'Table 5'.\n"
    "  2. The description's only substantive content is an attachment reference, "
    "     e.g. 'See Figure 3.', 'Refer to Table 12.', 'As shown in the schematic "
    "     below.', 'See attached diagram.' — and removing that reference leaves "
    "     nothing meaningful behind.\n\n"
    "Otherwise return type=\"text\". In particular, return type=\"text\" for:\n"
    "  - Short technical statements with concrete signal names, register names, or "
    "    conditions, even when only one short sentence. Examples: 'BIST is "
    "    performed', 'AT POR, if TEST_CFG_AT_POR=1 (00b), valid OTP configuration, "
    "    skip BIST at POR', 'The internal oscillator shall power up within 5us'.\n"
    "  - Brief one-liner behaviors, requirements, constraints, or section intros. "
    "    Brevity alone is NEVER figure_table.\n"
    "  - Items that mention a figure but ALSO contain meaningful prose — the prose "
    "    makes them text.\n\n"
    "When uncertain, default to type=\"text\". The cost of misclassifying a text "
    "item as figure_table is hiding a real coverage gap; the cost of the reverse "
    "is just one extra page-comparison call.\n\n"
    "Output ONLY a single JSON object — no preamble, no code fences. Start with { "
    "and end with }."
)

PRECLASS_USER = """Jama item:
ID: {jama_id}
Name: {name}
Description: {description}

Return JSON:
{{
  "type": "text" or "figure_table",
  "reason": "one short sentence"
}}
"""

CLASSIFY_SYSTEM = (
    "You are a requirements traceability analyst. You will be given a Jama "
    "requirement and several candidate PDF pages retrieved by similarity. Decide "
    "whether the PDF faithfully covers the requirement, and pick the single best "
    "matching page if any.\n\n"
    "Status definitions — BINARY (only two possible values):\n"
    "- identical: One of the candidate pages addresses the same substantive content "
    "as the Jama item. Different wording, synonyms, or paraphrasing is FINE. What "
    "matters is the technical content: same signal names, same behavior, same "
    "values. Example: Jama 'External pull-up', PDF 'pulled high externally' — "
    "IDENTICAL. In this case set matched_page to that page number and quoted_text "
    "to the exact substring from that page.\n"
    "- mismatch: Anything else. This covers BOTH of these situations:\n"
    "    (a) A candidate page discusses the same topic as the Jama item, but the "
    "        actual technical content differs — different numeric values, different "
    "        signal names, different logic, contradicting behavior. In this case "
    "        set matched_page to that page and quoted_text to the conflicting text "
    "        from that page.\n"
    "    (b) NONE of the candidate pages addresses the Jama item's topic. In this "
    "        case set matched_page to null and quoted_text to an empty string.\n\n"
    "When in doubt between identical and mismatch case (a), prefer identical. A "
    "wording difference alone is never a mismatch — only contradicting technical "
    "content is.\n\n"
    "Output ONLY a single JSON object — no preamble, no code fences. Start with { "
    "and end with }."
)

CLASSIFY_USER = """Jama item:
ID: {jama_id}
Name: {name}
Description: {description}

Candidate PDF pages (each page text is enclosed in triple quotes):
{pages_block}

Return JSON:
{{
  "status": "identical" | "mismatch",
  "matched_page": <page number from above, or null if no candidate page covers the topic>,
  "quoted_text": "<exact substring from the matched page that corresponds to the Jama item, or empty string if no page covers the topic>",
  "reasoning": "<2-3 sentences explaining the classification>"
}}
"""


def parse_llm_json(raw: str) -> dict:
    s = raw.strip()
    if s.startswith("```"):
        s = re.sub(r"^```(?:json)?\s*", "", s)
        s = re.sub(r"\s*```$", "", s)
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", s, re.DOTALL)
        if not m:
            raise
        return json.loads(m.group(0))


def llm_content(msg) -> str:
    """Return the message body. With qwen3 reasoning parser the JSON lands in `content`;
    `reasoning_content` is the chain-of-thought we want to ignore. We keep the fallback
    in case the server is started without the parser flag."""
    c = (msg.content or "").strip()
    if c:
        return c
    return (getattr(msg, "reasoning_content", None) or "").strip()


# ---------- LLM calls ----------

# Reasoning models occasionally burn the entire token budget on `<think>` and
# never emit the final JSON, so `content` comes back empty. We use a two-layer
# strategy:
#   Attempt 1: full reasoning enabled, generous max_tokens.
#   Attempt 2 (only on empty content): same prompt with `/no_think` appended,
#     which Qwen3 honors by skipping the <think> phase entirely. This guarantees
#     a non-empty response (no reasoning to overflow), at slightly lower quality
#     for those edge cases. Definite answer beats `error`.
NO_THINK_TOKEN = "\n\n/no_think"


def _disable_thinking(messages: list[dict]) -> list[dict]:
    """Return a copy of `messages` with `/no_think` appended to the last user
    message. Qwen3 reads this token from the user prompt and skips reasoning."""
    out = [dict(m) for m in messages]
    for m in reversed(out):
        if m.get("role") == "user":
            m["content"] = (m.get("content") or "") + NO_THINK_TOKEN
            break
    return out


async def _llm_call_with_fallback(
    client: AsyncOpenAI, model: str, messages: list[dict],
    max_tokens: int, fallback_max_tokens: int = 4000,
) -> tuple[str, dict]:
    """Run the chat completion. Returns (content, diagnostics).

    Two attempts:
      1. As-given, max_tokens=max_tokens, full reasoning.
      2. Only if attempt 1 returned empty content (not on exceptions): same prompt
         with /no_think appended, max_tokens=fallback_max_tokens.

    `diagnostics` is a dict capturing finish_reason and reasoning_content lengths
    for whichever attempts ran — written into the verdict's reasoning field so a
    reviewer can see why a row failed without re-running anything."""
    diag: dict = {}

    # Attempt 1
    try:
        resp = await client.chat.completions.create(
            model=model, messages=messages,
            temperature=0.0, max_tokens=max_tokens,
        )
        choice = resp.choices[0]
        content = llm_content(choice.message)
        diag["a1_finish"] = choice.finish_reason
        diag["a1_reason_len"] = len(getattr(choice.message, "reasoning_content", "") or "")
        if content:
            return content, diag
    except Exception as e:  # noqa: BLE001
        # Underlying call failed — surface, do not try the fallback (different
        # failure mode, retrying with /no_think likely fails the same way).
        raise

    # Attempt 2 — /no_think fallback. Only reached if attempt 1 returned empty.
    fb_messages = _disable_thinking(messages)
    try:
        resp = await client.chat.completions.create(
            model=model, messages=fb_messages,
            temperature=0.0, max_tokens=fallback_max_tokens,
        )
        choice = resp.choices[0]
        content = llm_content(choice.message)
        diag["a2_finish"] = choice.finish_reason
        diag["a2_reason_len"] = len(getattr(choice.message, "reasoning_content", "") or "")
        diag["a2_no_think"] = True
        if content:
            return content, diag
    except Exception as e:  # noqa: BLE001
        diag["a2_error"] = f"{type(e).__name__}: {e}"

    return "", diag


async def preclassify_one(client: AsyncOpenAI, model: str, req: Req) -> PreVerdict:
    user = PRECLASS_USER.format(
        jama_id=req.jama_id,
        name=req.name[:1000],
        description=req.description[:4000],
    )
    messages = [
        {"role": "system", "content": PRECLASS_SYSTEM},
        {"role": "user", "content": user},
    ]
    try:
        content, diag = await _llm_call_with_fallback(
            client, model, messages, max_tokens=12000, fallback_max_tokens=2000,
        )
    except Exception as e:  # noqa: BLE001
        return PreVerdict(row_index=req.row_index, is_text=True,
                          reason=f"LLM error — defaulting to text: {type(e).__name__}: {e}")
    if not content:
        return PreVerdict(row_index=req.row_index, is_text=True,
                          reason=f"empty after fallback — defaulting to text. diag={diag}")
    try:
        data = parse_llm_json(content)
    except json.JSONDecodeError:
        return PreVerdict(row_index=req.row_index, is_text=True,
                          reason=f"parse failed — defaulting to text. raw={content[:200]}",
                          raw=content[:500])
    t = str(data.get("type", "")).strip().lower()
    is_text = t != "figure_table"
    return PreVerdict(
        row_index=req.row_index,
        is_text=is_text,
        reason=str(data.get("reason", ""))[:500],
    )


async def classify_one(client: AsyncOpenAI, model: str, req: Req,
                       candidates: list[Page], page_text_limit: int) -> Verdict:
    pages_block_parts = []
    for p in candidates:
        text = p.text[:page_text_limit]
        pages_block_parts.append(f"[Page {p.page_number}]\n\"\"\"\n{text}\n\"\"\"")
    pages_block = "\n\n".join(pages_block_parts)

    user = CLASSIFY_USER.format(
        jama_id=req.jama_id,
        name=req.name[:1000],
        description=req.description[:4000],
        pages_block=pages_block,
    )
    messages = [
        {"role": "system", "content": CLASSIFY_SYSTEM},
        {"role": "user", "content": user},
    ]
    try:
        # 48k leaves the reasoning model lots of room to think; if it still hits
        # the cap (empty content), the fallback retries with /no_think to force
        # a direct answer without reasoning.
        content, diag = await _llm_call_with_fallback(
            client, model, messages, max_tokens=48000, fallback_max_tokens=4000,
        )
    except Exception as e:  # noqa: BLE001
        return Verdict(row_index=req.row_index, status=ST_ERROR,
                       matched_page=None, quoted_text="",
                       reasoning=f"LLM call failed: {type(e).__name__}: {e}")
    if not content:
        return Verdict(row_index=req.row_index, status=ST_ERROR,
                       matched_page=None, quoted_text="",
                       reasoning=f"LLM empty after /no_think fallback. diag={diag}")
    try:
        data = parse_llm_json(content)
    except json.JSONDecodeError:
        return Verdict(row_index=req.row_index, status=ST_ERROR,
                       matched_page=None, quoted_text="",
                       reasoning=f"JSON parse failed. raw={content[:500]}")
    status = str(data.get("status", "")).strip().lower()
    # Defensive: if the LLM (or a stale checkpoint) yields the legacy "missing"
    # value, coerce it into "mismatch". The customer's vocabulary is binary now.
    if status == "missing":
        status = ST_MISMATCH
    if status not in {ST_IDENTICAL, ST_MISMATCH}:
        return Verdict(row_index=req.row_index, status=ST_ERROR,
                       matched_page=None, quoted_text="",
                       reasoning=f"unexpected status from LLM: {status!r}")
    matched_raw = data.get("matched_page")
    try:
        matched_page = int(matched_raw) if matched_raw not in (None, "", "null") else None
    except (TypeError, ValueError):
        matched_page = None
    return Verdict(
        row_index=req.row_index,
        status=status,
        matched_page=matched_page,
        quoted_text=str(data.get("quoted_text", ""))[:4000],
        reasoning=str(data.get("reasoning", ""))[:1000],
    )


# ---------- checkpoints ----------

def load_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        return []
    out = []
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            out.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return out


def append_jsonl(path: Path, obj: dict) -> None:
    with path.open("a") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")


# ---------- runners ----------

async def run_pre(client, model, reqs_to_pre, checkpoint, done_keys, concurrency):
    sem = asyncio.Semaphore(concurrency)
    pbar = tqdm(total=len(reqs_to_pre),
                initial=sum(1 for r in reqs_to_pre if r.row_index in done_keys),
                desc="pre-classify", unit="req")
    results: list[PreVerdict] = []

    async def worker(req: Req):
        if req.row_index in done_keys:
            return
        async with sem:
            v = await preclassify_one(client, model, req)
        append_jsonl(checkpoint, asdict(v))
        results.append(v)
        pbar.update(1)

    await asyncio.gather(*(worker(r) for r in reqs_to_pre))
    pbar.close()
    return results


async def run_classify(client, model, jobs, checkpoint, done_keys,
                       concurrency, page_text_limit):
    sem = asyncio.Semaphore(concurrency)
    pbar = tqdm(total=len(jobs),
                initial=sum(1 for r, _ in jobs if r.row_index in done_keys),
                desc="classify", unit="req")
    results: list[Verdict] = []

    async def worker(req: Req, candidates: list[Page]):
        if req.row_index in done_keys:
            return
        async with sem:
            v = await classify_one(client, model, req, candidates, page_text_limit)
        append_jsonl(checkpoint, asdict(v))
        results.append(v)
        pbar.update(1)

    await asyncio.gather(*(worker(r, c) for r, c in jobs))
    pbar.close()
    return results


# ---------- writer ----------

def write_template(input_xlsx: Path, output_xlsx: Path,
                   reqs: list[Req], pre_by_row: dict[int, PreVerdict],
                   verdicts_by_row: dict[int, Verdict]) -> None:
    """Open input xlsx, fill PDF Text Content / Status / Key Difference per row,
    save as output_xlsx. Column A formulas (HYPERLINK) are preserved."""
    shutil.copyfile(input_xlsx, output_xlsx)
    wb = load_workbook(output_xlsx, data_only=False)
    ws = wb.worksheets[0]  # Sheet1

    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value).strip() if cell.value is not None else "")

    pdf_text_idx = find_header_col(headers, PDF_TEXT_COL_CANDIDATES)
    status_idx = find_header_col(headers, STATUS_COL_CANDIDATES)
    keydiff_idx = find_header_col(headers, KEY_DIFF_COL_CANDIDATES)

    missing = []
    if pdf_text_idx < 0:
        missing.append("PDF Text Content")
    if status_idx < 0:
        missing.append("Status")
    if keydiff_idx < 0:
        missing.append("Key Difference")
    if missing:
        raise SystemExit(
            f"could not locate columns in input xlsx: {missing}. "
            f"headers found: {headers}"
        )

    pdf_text_col = pdf_text_idx + 1
    status_col = status_idx + 1
    keydiff_col = keydiff_idx + 1

    print(f"writing into columns: PDF Text={get_column_letter(pdf_text_col)}, "
          f"Status={get_column_letter(status_col)}, "
          f"Key Difference={get_column_letter(keydiff_col)}")

    wrap = Alignment(wrap_text=True, vertical="top")

    counts: dict[str, int] = {}
    for req in reqs:
        # Skip Folder/Heading/etc.: write Item Type into Status, leave D and F blank.
        if req.is_skip_type:
            label = req.skip_status or "Folder"
            ws.cell(row=req.sheet_row, column=status_col, value=label).fill = SKIP_TYPE_FILL
            ws.cell(row=req.sheet_row, column=status_col).alignment = wrap
            counts[label] = counts.get(label, 0) + 1
            continue

        pre = pre_by_row.get(req.row_index)
        if pre and not pre.is_text:
            ws.cell(row=req.sheet_row, column=pdf_text_col, value="").alignment = wrap
            sc = ws.cell(row=req.sheet_row, column=status_col, value=ST_FIGURE)
            sc.fill = STATUS_FILLS[ST_FIGURE]
            sc.alignment = wrap
            ws.cell(row=req.sheet_row, column=keydiff_col,
                    value=pre.reason).alignment = wrap
            counts[ST_FIGURE] = counts.get(ST_FIGURE, 0) + 1
            continue

        v = verdicts_by_row.get(req.row_index)
        if v is None:
            ws.cell(row=req.sheet_row, column=status_col, value="").alignment = wrap
            counts["unprocessed"] = counts.get("unprocessed", 0) + 1
            continue

        # PDF Text Content cell — quote + (p.X) suffix.
        pdf_text_value = ""
        if v.quoted_text and v.matched_page is not None:
            pdf_text_value = f"{v.quoted_text}\n\n(p.{v.matched_page})"
        elif v.quoted_text:
            pdf_text_value = v.quoted_text
        ws.cell(row=req.sheet_row, column=pdf_text_col, value=pdf_text_value).alignment = wrap

        # Status cell with color fill.
        sc = ws.cell(row=req.sheet_row, column=status_col, value=v.status)
        fill = STATUS_FILLS.get(v.status)
        if fill:
            sc.fill = fill
        sc.alignment = wrap

        # Key Difference cell. For mismatch we want the word-diff (when there's a
        # quoted PDF passage to compare against) plus the model's reasoning. When
        # quoted_text is empty (the "absent from PDF" subtype of mismatch) the
        # diff trivially returns "" and we just show the reasoning.
        if v.status == ST_IDENTICAL:
            diff = word_diff(req.description, v.quoted_text)
            kd = diff if diff else "(no textual differences)"
        elif v.status == ST_MISMATCH:
            diff = word_diff(req.description, v.quoted_text)
            kd = (diff + "\n\n" if diff else "") + v.reasoning
        else:
            kd = v.reasoning
        ws.cell(row=req.sheet_row, column=keydiff_col, value=kd[:1500]).alignment = wrap

        counts[v.status] = counts.get(v.status, 0) + 1

    wb.save(output_xlsx)

    print("\nstatus counts:")
    for k in sorted(counts.keys()):
        print(f"  {k:<20} {counts[k]}")


# ---------- main ----------

def main() -> int:
    p = argparse.ArgumentParser(description="run2 — fill the customer's Excel template "
                                            "with PDF coverage results.")
    p.add_argument("xlsx", type=Path, help="input Jama xlsx")
    p.add_argument("pdf", type=Path, help="input PDF (slice is fine; pages report local 1..N)")
    p.add_argument("--out", type=Path, default=None,
                   help="output xlsx path (default: <input>_filled.xlsx)")
    p.add_argument("--limit", type=int, default=None,
                   help="process only the first N requirement rows for a smoke test")
    p.add_argument("--re-extract", action="store_true",
                   help="force re-extracting excel.json / pdf.json from the inputs")
    p.add_argument("--top-k", type=int, default=DEFAULT_TOP_K)
    p.add_argument("--concurrency", type=int, default=DEFAULT_CONCURRENCY)
    p.add_argument("--llm-url", default=DEFAULT_LLM_URL)
    p.add_argument("--llm-model", default=DEFAULT_LLM_MODEL)
    p.add_argument("--embed-model", default=DEFAULT_EMBED_MODEL)
    p.add_argument("--page-text-limit", type=int, default=DEFAULT_PAGE_TEXT_LIMIT)
    p.add_argument("--pre-checkpoint", type=Path, default=Path("preclassify.jsonl"))
    p.add_argument("--cls-checkpoint", type=Path, default=Path("classify.jsonl"))
    args = p.parse_args()

    if not args.xlsx.exists():
        print(f"ERR: xlsx not found: {args.xlsx}", file=sys.stderr)
        return 1
    if not args.pdf.exists():
        print(f"ERR: pdf not found: {args.pdf}", file=sys.stderr)
        return 1

    work_dir = args.xlsx.parent
    out_path = args.out or (work_dir / f"{args.xlsx.stem}_filled.xlsx")

    excel_json, pdf_json = ensure_extracted(args.xlsx, args.pdf, work_dir, args.re_extract)

    reqs, meta, columns = load_reqs(excel_json)
    pages = load_pages(pdf_json)
    print(f"loaded {len(reqs)} rows | {len(pages)} pages | column keys: {meta}")

    # Drop blank trailing rows (no Jama ID) — they otherwise go through the full
    # pipeline with degenerate input and tend to stall the run. Excel readers
    # often surface a few empty tail rows after the real data ends. We leave
    # those rows untouched in the output xlsx (D/E/F stay blank).
    pre_filter_count = len(reqs)
    reqs = [r for r in reqs if r.jama_id]
    dropped_no_id = pre_filter_count - len(reqs)
    if dropped_no_id:
        print(f"dropped {dropped_no_id} rows with no Jama ID (left blank in output)")

    if args.limit:
        reqs = reqs[:args.limit]
        print(f"--limit {args.limit}: processing first {len(reqs)} rows")

    if not reqs or not pages:
        print("nothing to do.", file=sys.stderr)
        return 1

    # Reqs we need to pre-classify (skip-type rows bypass everything).
    real_reqs = [r for r in reqs if not r.is_skip_type]
    print(f"non-skip-type rows: {len(real_reqs)}")

    pre_ckpt_path = work_dir / args.pre_checkpoint
    cls_ckpt_path = work_dir / args.cls_checkpoint
    pre_done_records = load_jsonl(pre_ckpt_path)
    cls_done_records = load_jsonl(cls_ckpt_path)
    pre_done_keys = {d["row_index"] for d in pre_done_records}
    cls_done_keys = {d["row_index"] for d in cls_done_records}
    print(f"pre checkpoint: {len(pre_done_records)} records | "
          f"classify checkpoint: {len(cls_done_records)} records")

    # HTTP/LLM client. trust_env=False prevents corp HTTP_PROXY from hijacking localhost.
    http_client = httpx.AsyncClient(trust_env=False, timeout=httpx.Timeout(DEFAULT_LLM_TIMEOUT))
    client = AsyncOpenAI(base_url=args.llm_url, api_key="sk-local", http_client=http_client)

    # ---- pre-classify ----
    print(f"\n>>> pre-classifying {len(real_reqs)} non-skip rows...")
    fresh_pre = asyncio.run(run_pre(
        client, args.llm_model, real_reqs, pre_ckpt_path, pre_done_keys, args.concurrency,
    ))
    pre_records = pre_done_records + [asdict(v) for v in fresh_pre]
    pre_by_row: dict[int, PreVerdict] = {}
    for d in pre_records:
        pre_by_row[d["row_index"]] = PreVerdict(
            row_index=d["row_index"],
            is_text=bool(d.get("is_text", True)),
            reason=str(d.get("reason", "")),
            raw=str(d.get("raw", "") or ""),
        )

    text_reqs = [r for r in real_reqs if pre_by_row.get(r.row_index) and pre_by_row[r.row_index].is_text]
    figure_count = len(real_reqs) - len(text_reqs)
    print(f"text rows: {len(text_reqs)} | figure/table rows: {figure_count}")

    # ---- embed + retrieve ----
    if not text_reqs:
        print("no text rows to classify; writing template now.")
        write_template(args.xlsx, out_path, reqs, pre_by_row, {})
        print(f"\nreport written -> {out_path}")
        return 0

    print(f"\n>>> embedding with {args.embed_model}...")
    embed_model = SentenceTransformer(args.embed_model, device="cpu")
    req_vecs = embed(embed_model, [r.embed_text for r in text_reqs], "reqs")
    page_vecs = embed(embed_model, [p.text for p in pages], "pages")
    sim = req_vecs @ page_vecs.T  # (R, P), normalized

    jobs: list[tuple[Req, list[Page]]] = []
    for i, req in enumerate(text_reqs):
        idxs = top_k_indices(sim[i], args.top_k)
        candidates = [pages[j] for j in idxs]
        jobs.append((req, candidates))
    print(f"classify jobs: {len(jobs)} (top-{args.top_k} pages each)")

    # ---- classify ----
    print(f"\n>>> classifying with {args.llm_model}...")
    fresh_cls = asyncio.run(run_classify(
        client, args.llm_model, jobs, cls_ckpt_path, cls_done_keys,
        args.concurrency, args.page_text_limit,
    ))
    cls_records = cls_done_records + [asdict(v) for v in fresh_cls]
    verdicts_by_row: dict[int, Verdict] = {}
    for d in cls_records:
        # Coerce legacy "missing" status (from earlier runs) into "mismatch" so
        # stale checkpoint records render correctly with the new binary scheme.
        loaded_status = str(d.get("status", ""))
        if loaded_status == "missing":
            loaded_status = ST_MISMATCH
        verdicts_by_row[d["row_index"]] = Verdict(
            row_index=d["row_index"],
            status=loaded_status,
            matched_page=d.get("matched_page"),
            quoted_text=str(d.get("quoted_text", "")),
            reasoning=str(d.get("reasoning", "")),
        )

    # ---- write ----
    write_template(args.xlsx, out_path, reqs, pre_by_row, verdicts_by_row)
    print(f"\nreport written -> {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
